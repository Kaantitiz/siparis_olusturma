import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from functools import lru_cache
import re
from difflib import SequenceMatcher

# Cache temizleme fonksiyonu
def clear_all_caches():
    """Tüm cache'leri temizle"""
    try:
        # Cache temizleme
        st.cache_data.clear()
        st.cache_resource.clear()
        
        # Session state temizleme
        if 'processed_data' in st.session_state:
            del st.session_state.processed_data
        if 'brand_data_cache' in st.session_state:
            del st.session_state.brand_data_cache
        
        return True
    except Exception as e:
        st.error(f"Cache temizleme hatası: {str(e)}")
        return False

# Ürün kodu eşleştirme yardımcı fonksiyonları
def clean_product_code(code):
    """Ürün kodunu temizle ve standardize et"""
    if pd.isna(code) or code == '':
        return ''
    
    # String'e çevir
    code_str = str(code).strip()
    
    # Boşlukları kaldır
    code_str = code_str.replace(' ', '').replace('-', '').replace('_', '')
    
    # Büyük harfe çevir
    code_str = code_str.upper()
    
    # Özel karakterleri temizle (sadece harf, rakam ve nokta bırak)
    import re
    code_str = re.sub(r'[^A-Z0-9.]', '', code_str)
    
    return code_str

def find_best_match(product_code, target_codes, threshold=0.8):
    """En iyi eşleşmeyi bul (fuzzy matching)"""
    from difflib import SequenceMatcher
    
    if not product_code:
        return None, 0
    
    best_match = None
    best_ratio = 0
    
    for target_code in target_codes:
        if pd.isna(target_code):
            continue
            
        target_str = str(target_code).strip()
        
        # Tam eşleşme kontrolü
        if clean_product_code(product_code) == clean_product_code(target_str):
            return target_code, 1.0
        
        # Fuzzy matching
        ratio = SequenceMatcher(None, clean_product_code(product_code), clean_product_code(target_str)).ratio()
        
        if ratio > best_ratio and ratio >= threshold:
            best_ratio = ratio
            best_match = target_code
    
    return best_match, best_ratio

def process_schaeffler_codes(catalogue_number):
    """Schaeffler ürün kodlarını işle"""
    if pd.isna(catalogue_number):
        return ''
    
    code_str = str(catalogue_number).strip()
    
    # Özel Schaeffler kuralları
    # 1. Sondaki 0'ları kaldır (sadece belirli durumlarda)
    if code_str.endswith('0') and len(code_str) > 1:
        # Eğer sondaki 0'dan önceki karakter rakam değilse, 0'ı kaldır
        if not code_str[-2].isdigit():
            code_str = code_str[:-1]
    
    # 2. Özel Schaeffler formatları
    # LUK formatı: LUK-XXXXX -> XXXXX
    if code_str.startswith('LUK-'):
        code_str = code_str[4:]
    
    # 3. Boşlukları ve özel karakterleri temizle
    code_str = clean_product_code(code_str)
    
    return code_str

def process_valeo_codes(valeo_ref):
    """Valeo ürün kodlarını işle"""
    if pd.isna(valeo_ref):
        return ''
    
    code_str = str(valeo_ref).strip()
    
    # Özel Valeo kuralları
    # 1. Valeo özel formatları
    # VALE-XXXXX -> XXXXX
    if code_str.startswith('VALE-'):
        code_str = code_str[5:]
    
    # 2. Boşlukları ve özel karakterleri temizle
    code_str = clean_product_code(code_str)
    
    return code_str

# Sayfa ayarları
st.set_page_config(
    page_title="Excel Dönüştürme Aracı (Ultra Hızlı)",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Başlık
st.title("⚡ Ultra Hızlı Excel Dönüştürücü")
st.caption("100.000+ satırlık dosyalar için optimize edilmiş versiyon - Maksimum Hız Modu")

# Schaeffler ve Valeo ürün kodu eşleştirme iyileştirmeleri
st.info("🔧 **Geliştirmeler:** Schaeffler ve Valeo ürün kodu eşleştirme algoritması iyileştirildi!")
st.write("""
**Yeni Özellikler:**
- ✅ Geliştirilmiş ürün kodu temizleme
- ✅ Fuzzy matching (benzer kodlar için)
- ✅ Hem URUNKODU hem Düzenlenmiş Ürün Kodu ile eşleştirme
- ✅ Detaylı debug bilgileri
- ✅ Daha iyi hata yönetimi
""")

# Global değişkenler
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'brand_data_cache' not in st.session_state:
    st.session_state.brand_data_cache = {}
if 'app_restart_count' not in st.session_state:
    st.session_state.app_restart_count = 0

# Ultra hızlı önbellek fonksiyonları
@st.cache_data(max_entries=5, show_spinner="Dosya okunuyor...", ttl=3600)
def load_data_ultra_fast(uploaded_file):
    """Maksimum hızlı dosya okuma"""
    try:
        # Maksimum hız için minimal ayarlar
        df = pd.read_excel(
            uploaded_file,
            engine='openpyxl',
            # dtype belirtme - sadece kritik sütunlar
            dtype={
                'URUNKODU': 'string'
            },
            # NaN kontrolü tamamen devre dışı
            na_filter=False,
            keep_default_na=False,
            # Ek hızlandırma
            header=0,
            skiprows=None,
            nrows=None  # Tüm satırları oku
        )
        
        return df
    except Exception as e:
        st.error(f"Dosya okuma hatası: {str(e)}")
        return pd.DataFrame()

@st.cache_data(show_spinner="Marka verisi okunuyor...", ttl=1800)
def load_brand_data_parallel(excel_file, brand_name):
    """Maksimum hızlı marka verisi okuma"""
    try:
        # Maksimum hız için minimal ayarlar
        df = pd.read_excel(
            excel_file,
            engine='openpyxl',
            na_filter=False,
            keep_default_na=False
        )
        
        return brand_name, df
    except Exception as e:
        return brand_name, pd.DataFrame()

@st.cache_data(show_spinner="Veri dönüştürülüyor...", ttl=3600)
def transform_data_ultra_fast(df):
    """Maksimum hızlı veri dönüştürme"""
    try:
        # Sadece gerekli sütunları al - bellek tasarrufu
        essential_cols = [
            'URUNKODU', 'ACIKLAMA', 'URETİCİKODU', 'ORJİNAL', 'ESKİKOD',
            'TOPL.FAT.ADT', 'MÜŞT.SAY.', 'SATıŞ FIYATı', 'DÖVIZ CINSI (S)'
        ] + [f'CAT{i}' for i in range(1, 8)]
        
        # Depo sütunları - sadece mevcut olanları al
        depo_prefixes = ['02-', '04-', 'D01-', 'A01-', 'TD-E01-', 'E01-']
        depo_cols = []
        for prefix in depo_prefixes:
            for col_type in ['DEVIR', 'ALIS', 'STOK', 'SATIS']:
                col_name = f"{prefix}{col_type}"
                if col_name in df.columns:
                    depo_cols.append(col_name)
        
        # Mevcut sütunları filtrele
        available_cols = [col for col in essential_cols + depo_cols if col in df.columns]
        df_filtered = df[available_cols].copy()
        
        # Maksimum hızlı dönüşüm - vektörel işlemler
        new_df = pd.DataFrame()
        
        # 1. URUNKODU (ilk) - vektörel
        new_df['URUNKODU'] = df_filtered['URUNKODU'].fillna(0)
        
        # 2. Düzenlenmiş Ürün Kodu - vektörel (başında 0 olan kodlar için özel format)
        new_df['Düzenlenmiş Ürün Kodu'] = df_filtered['URUNKODU'].fillna(0).str.replace(r'^[^-]*-', "", regex=True)
        
        # 4-7. Temel sütunlar - vektörel
        basic_cols = ['ACIKLAMA', 'URETİCİKODU', 'ORJİNAL', 'ESKİKOD']
        for col in basic_cols:
            if col in df_filtered.columns:
                new_df[col] = df_filtered[col].fillna(0)
        
        # 8. Kategoriler - vektörel
        for i in range(1, 8):
            cat_col = f'CAT{i}'
            if cat_col in df_filtered.columns:
                new_df[f'CAT{i}'] = df_filtered[cat_col].fillna(0)
        
        # 9. Depo verileri - vektörel işlem
        depo_mapping = {
            '02-': 'MASLAK',
            'D01-': 'İMES',
            'TD-E01-': 'İKİTELLİ',
            'E01-': 'İKİTELLİ',
            '04-': 'BOLU',
            'A01-': 'ANKARA'
        }
        
        for old_prefix, new_name in depo_mapping.items():
            for col_type, new_type in zip(['DEVIR', 'ALIS', 'SATIS', 'STOK'],
                                         ['DEVIR', 'ALIŞ', 'SATIS', 'STOK']):
                old_col = f"{old_prefix}{col_type}"
                if old_col in df_filtered.columns:
                    # Vektörel işlem
                    col_data = df_filtered[old_col].fillna(0)
                    if pd.api.types.is_numeric_dtype(col_data):
                        col_data = col_data.astype(float)
                        col_data = col_data.replace(0, '-')
                    else:
                        col_data = col_data.astype(str)
                    new_df[f"{new_name} {new_type}"] = col_data.astype('string')
                else:
                    # Eksik sütun için boş değer
                    new_df[f"{new_name} {new_type}"] = '-'
        
        # 10. Tedarikçi bakiye kolonları - vektörel
        tedarikci_cols = [
            'İmes Tedarikçi Bakiye', 'Ankara Tedarikçi Bakiye', 
            'Bolu Tedarikçi Bakiye', 'Maslak Tedarikçi Bakiye'
        ]
        
        for col in tedarikci_cols:
            new_df[col] = '-'
        
        # 11. Dinamik ay başlıkları - vektörel
        current_month = datetime.datetime.now().month
        months = ['Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
                 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
        
        next_month1 = months[(current_month) % 12]
        next_month2 = months[(current_month + 1) % 12]
        
        # Vektörel ay başlıkları
        for i in range(5):
            new_df[f'{next_month1}_{i+1}'] = 0
            new_df[f'{next_month2}_{i+1}'] = 0
        
        # 12. Diğer sütunlar - vektörel
        other_cols = {
            'TOPL.FAT.ADT': 'TOPL.FAT.ADT',
            'MÜŞT.SAY.': 'MÜŞT.SAY.',
            'SATıŞ FIYATı': 'SATıŞ FIYATı',
            'DÖVIZ CINSI (S)': 'DÖVIZ CINSI (S)'
        }
        
        for old, new in other_cols.items():
            if old in df_filtered.columns:
                new_df[new] = df_filtered[old].fillna(0)
        
        # 13. URUNKODU (DÖVIZ CINSI'den sonra)
        new_df['URUNKODU_3'] = df_filtered['URUNKODU'].fillna(0)
        
        # 14. Eksik başlıkları geri getir - vektörel
        # not, İSK, PRİM, BÜTÇE, liste, TD SF, Net Fiyat Kampanyası
        new_df['not'] = 0
        new_df['İSK'] = 0
        new_df['PRİM'] = 0
        new_df['BÜTÇE'] = 0
        new_df['liste'] = 0
        new_df['TD SF'] = 0
        new_df['Net Fiyat Kampanyası'] = 0
        
        # Kampanya Tipi
        new_df['Kampanya Tipi'] = 0
        
        # Toplam İsk
        new_df['Toplam İsk'] = 0
        
        # Depo Bakiye kolonları
        new_df['Maslak Depo Bakiye'] = 0
        new_df['Bolu Depo Bakiye'] = 0
        new_df['İmes Depo Bakiye'] = 0
        new_df['Ankara Depo Bakiye'] = 0
        new_df['İkitelli Depo Bakiye'] = 0
        
        # Toplam Depo Bakiye - otomatik hesaplama
        new_df['Toplam Depo Bakiye'] = 0
        
        # Tedarikçi bakiye kolonları - İkitelli Tedarikçi Bakiye eklendi
        tedarikci_cols = [
            'İmes Tedarikçi Bakiye', 'Ankara Tedarikçi Bakiye', 
            'Bolu Tedarikçi Bakiye', 'Maslak Tedarikçi Bakiye', 'İkitelli Tedarikçi Bakiye'
        ]
        
        for col in tedarikci_cols:
            new_df[col] = 0
        
        # Paket Adetleri
        new_df['Paket Adetleri'] = 0
        
        # Sipariş kolonları
        new_df['Maslak Sipariş'] = 0
        new_df['Bolu Sipariş'] = 0
        new_df['İmes Sipariş'] = 0
        new_df['Ankara Sipariş'] = 0
        new_df['İkitelli Sipariş'] = 0
        
        # Sütun sıralamasını düzelt - verilen sıraya göre (64 adet)
        desired_order = [
            'URUNKODU', 'Düzenlenmiş Ürün Kodu', 'ACIKLAMA', 'URETİCİKODU', 'ORJİNAL', 'ESKİKOD',
            'CAT1', 'CAT2', 'CAT3', 'CAT4', 'CAT5', 'CAT6', 'CAT7',
            # Depo kolonları (sıralama: MASLAK, İMES, İKİTELLİ, BOLU, ANKARA)
            'MASLAK DEVIR', 'MASLAK ALIŞ', 'MASLAK SATIS', 'MASLAK STOK',
            'İMES DEVIR', 'İMES ALIŞ', 'İMES SATIS', 'İMES STOK',
            'İKİTELLİ DEVIR', 'İKİTELLİ ALIŞ', 'İKİTELLİ SATIS', 'İKİTELLİ STOK',
            'BOLU DEVIR', 'BOLU ALIŞ', 'BOLU SATIS', 'BOLU STOK',
            'ANKARA DEVIR', 'ANKARA ALIŞ', 'ANKARA SATIS', 'ANKARA STOK',
            # no2
            'not',
            # Depo Bakiye kolonları
            'Maslak Depo Bakiye', 'Bolu Depo Bakiye', 'İmes Depo Bakiye', 'Ankara Depo Bakiye', 'İkitelli Depo Bakiye',
            # Kampanya Tipi
            'Kampanya Tipi',
            # Toplam İsk
            'Toplam İsk',
            # Toplam Depo Bakiye
            'Toplam Depo Bakiye',
            # Tedarikçi bakiye kolonları
            'Maslak Tedarikçi Bakiye', 'Bolu Tedarikçi Bakiye', 'İmes Tedarikçi Bakiye', 'Ankara Tedarikçi Bakiye', 'İkitelli Tedarikçi Bakiye',
            # Paket Adetleri
            'Paket Adetleri',
            # Sipariş kolonları
            'Maslak Sipariş', 'Bolu Sipariş', 'İmes Sipariş', 'Ankara Sipariş', 'İkitelli Sipariş',
            # Ay başlıkları
            'Ağustos_1', 'Eylül_1', 'Ağustos_2', 'Eylül_2', 'Ağustos_3', 'Eylül_3', 'Ağustos_4', 'Eylül_4', 'Ağustos_5', 'Eylül_5',
            # Diğer sütunlar
            'TOPL.FAT.ADT', 'MÜŞT.SAY.', 'SATıŞ FIYATı', 'DÖVIZ CINSI (S)', 'URUNKODU_3',
            # Son başlıklar
            'Kampanya Tipi', 'not', 'İSK', 'PRİM', 'BÜTÇE', 'liste', 'TD SF', 'Toplam İsk', 'Net Fiyat Kampanyası'
        ]
        
        # Mevcut sütunları filtrele ve sırala
        available_cols = [col for col in desired_order if col in new_df.columns]
        if len(available_cols) > 0:
            new_df = new_df[available_cols]
        
        # Toplam Depo Bakiye hesaplama
        depo_bakiye_cols = ['Maslak Depo Bakiye', 'Bolu Depo Bakiye', 'İmes Depo Bakiye', 'Ankara Depo Bakiye', 'İkitelli Depo Bakiye']
        available_depo_cols = [col for col in depo_bakiye_cols if col in new_df.columns]
        
        if available_depo_cols and 'Toplam Depo Bakiye' in new_df.columns:
            # Sayısal değerlere çevir ve topla
            for col in available_depo_cols:
                new_df[col] = pd.to_numeric(new_df[col], errors='coerce').fillna(0)
            
            # Toplam hesapla
            new_df['Toplam Depo Bakiye'] = new_df[available_depo_cols].sum(axis=1)
        
        return new_df
    
    except Exception as e:
        st.error(f"Dönüşüm hatası: {str(e)}")
        return pd.DataFrame()

@st.cache_data(show_spinner="Marka eşleştirme yapılıyor...", ttl=3600)
def match_brands_parallel(main_df, uploaded_files):
    """Paralel marka eşleştirme"""
    try:
        # Marka-Excel eşleştirme sözlüğü
        brand_excel_mapping = {
            'SCHAEFFLER LUK': 'excel1',
            'SCHAFLERR': 'excel1',  # Schaflerr için alternatif isim
            'ZF İTHAL': 'excel2', 
            'DELPHI': 'excel3',
            'ZF YERLİ': 'excel4',
            'VALEO': 'excel5',
            'FILTRON': 'excel6',
            'MANN': 'excel7'
        }
        
        # Ana DataFrame'i kopyala
        result_df = main_df.copy()
        
        # CAT4 kolonunu kontrol et
        if 'CAT4' not in main_df.columns:
            st.warning("CAT4 kolonu bulunamadı!")
            return main_df
        
        # Paralel işleme için marka verilerini topla
        brand_tasks = []
        for brand, excel_key in brand_excel_mapping.items():
            if excel_key in uploaded_files and uploaded_files[excel_key] is not None:
                brand_tasks.append((brand, uploaded_files[excel_key]))
        
        # Paralel marka verisi okuma
        brand_data = {}
        with ThreadPoolExecutor(max_workers=4) as executor:
            future_to_brand = {
                executor.submit(load_brand_data_parallel, file, brand): brand 
                for brand, file in brand_tasks
            }
            
            for future in as_completed(future_to_brand):
                brand_name, brand_df = future.result()
                brand_data[brand_name] = brand_df
                st.success(f"✅ {brand_name} verisi yüklendi: {len(brand_df)} satır")
        
        # Her marka için işlem yap
        for brand, brand_df in brand_data.items():
            if len(brand_df) > 0:
                # CAT4'te bu markayı ara (esnek arama)
                search_terms = [brand]
                
                # Schaeffler için özel arama terimleri - CAT4'teki tam değere göre
                if 'Schaeffler' in brand or 'Schaflerr' in brand:
                    search_terms = ['SCHAEFFLER LUK']  # CAT4'teki tam değer
                
                # Delphi için özel arama terimleri - CAT4'teki tam değere göre
                if 'DELPHI' in brand:
                    search_terms = ['DELPHI']  # CAT4'teki tam değer
                
                # ZF için özel arama terimleri
                if 'ZF' in brand:
                    search_terms.extend(['LEMFÖRDER', 'TRW', 'SACHS', 'LEMFORDER', 'TRW', 'SACHS'])
                
                # Mann için özel arama terimleri
                if 'MANN' in brand:
                    search_terms.extend(['MANN', 'MANN FILTER', 'MANN-FILTER', 'MANNFILTER'])
                
                # Filtron için özel arama terimleri
                if 'FILTRON' in brand:
                    search_terms.extend(['FILTRON'])
                
                # Debug: Arama terimlerini göster
                st.info(f"🔍 {brand} için arama terimleri: {search_terms}")
                
                # Tüm arama terimlerini dene
                brand_mask = pd.Series([False] * len(main_df))
                for search_term in search_terms:
                    temp_mask = main_df['CAT4'].str.contains(search_term, case=False, na=False)
                    brand_mask = brand_mask | temp_mask
                
                brand_count = brand_mask.sum()
                
                # Debug: CAT4'teki benzersiz değerleri göster
                if brand_count == 0:
                    unique_cat4 = main_df['CAT4'].dropna().unique()
                    st.info(f"🔍 CAT4 kolonundaki benzersiz değerler: {list(unique_cat4[:10])}")
                    st.warning(f"⚠️ {brand} markası CAT4 kolonunda bulunamadı! Lütfen CAT4 kolonundaki marka isimlerini kontrol edin.")
                    
                    # CAT4'te tam eşleşme ara
                    exact_matches = main_df[main_df['CAT4'] == search_terms[0]]
                    if len(exact_matches) > 0:
                        st.success(f"✅ Tam eşleşme bulundu: {search_terms[0]} - {len(exact_matches)} satır")
                        brand_mask = main_df['CAT4'] == search_terms[0]
                        brand_count = brand_mask.sum()
                else:
                    st.success(f"✅ {brand} markası {brand_count} ürün için bulundu")
                    
                    # Debug: Bulunan ürünlerin örneklerini göster
                    found_products = main_df[brand_mask]['CAT4'].head(5).tolist()
                    st.info(f"📋 Bulunan ürün örnekleri: {found_products}")
                    
                    # Mann ve Filtron için normal işlem (CAT4'te bulundu)
                    if ('MANN' in brand or 'FILTRON' in brand) and brand_count > 0:
                        st.info(f"🔄 {brand} için normal işlem yapılıyor (CAT4'te bulundu)...")
                        # Burada normal işlem yapılacak (Schaeffler gibi)
                        # Şimdilik boş bırakıyoruz, özel işlem kısmında yapılacak
                    
                    # Delphi ve Schaflerr için işlem yapılması gereken koşul
                    if ('DELPHI' in brand or 'SCHAEFFLER LUK' in brand or 'SCHAFLERR' in brand) and brand_count > 0:
                        st.info(f"🔄 {brand} için tedarikçi bakiye işlemi yapılıyor (CAT4'te bulundu)...")
                    
                    # Schaeffler Luk için tedarikçi bakiye işlemi
                    if 'SCHAEFFLER LUK' in brand or 'SCHAFLERR' in brand:
                        try:
                            # Tedarikçi bakiye kolonlarını oluştur
                            if 'İmes Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İmes Tedarikçi Bakiye'] = 0
                            if 'Ankara Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarikçi Bakiye'] = 0
                            if 'Bolu Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarikçi Bakiye'] = 0
                            if 'Maslak Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarikçi Bakiye'] = 0
                            if 'İkitelli Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İkitelli Tedarikçi Bakiye'] = 0
                            
                            # Schaeffler verilerini işle
                            schaeffler_df = brand_df.copy()
                            
                            # PO Number(L) kolonunu kontrol et
                            if 'PO Number(L)' in schaeffler_df.columns:
                                # Tedarikçi kodlarını belirle
                                schaeffler_df['Tedarikçi'] = schaeffler_df['PO Number(L)'].astype(str).apply(
                                    lambda x: 'İmes' if 'IME' in x or '285' in x
                                    else 'Ankara' if 'ANK' in x or '321' in x
                                    else 'Bolu' if '322' in x
                                    else 'Maslak' if '323' in x
                                    else 'İkitelli' if 'IKI' in x or '324' in x
                                    else 'Diğer'
                                )
                                
                                # Catalogue Number işleme - Geliştirilmiş
                                if 'Catalogue number' in schaeffler_df.columns:
                                    # Geliştirilmiş Schaeffler kod işleme
                                    schaeffler_df['Catalogue_clean'] = schaeffler_df['Catalogue number'].apply(process_schaeffler_codes)
                                    
                                    # Debug: Catalogue number işleme örnekleri göster
                                    st.info(f"🔍 Schaeffler Catalogue number işleme örnekleri (Geliştirilmiş):")
                                    for i, (original, cleaned) in enumerate(zip(schaeffler_df['Catalogue number'].head(), schaeffler_df['Catalogue_clean'].head())):
                                        st.write(f"  {original} → {cleaned}")
                                    
                                    # Eşleştirme istatistikleri
                                    total_codes = len(schaeffler_df['Catalogue_clean'])
                                    unique_codes = schaeffler_df['Catalogue_clean'].nunique()
                                    st.info(f"📊 Schaeffler kod istatistikleri: Toplam {total_codes}, Benzersiz {unique_codes}")
                                
                                # Ordered Quantity kontrolü
                                if 'Ordered quantity' in schaeffler_df.columns:
                                    # Tedarikçi bazında grupla ve topla
                                    for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                        tedarikci_data = schaeffler_df[schaeffler_df['Tedarikçi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Catalogue number bazında topla
                                            grouped = tedarikci_data.groupby('Catalogue_clean')['Ordered quantity'].sum().reset_index()
                                            
                                            # Ana DataFrame ile eşleştir - Geliştirilmiş
                                            for _, row in grouped.iterrows():
                                                catalogue_num = row['Catalogue_clean']
                                                quantity = row['Ordered quantity']
                                                
                                                # Geliştirilmiş eşleştirme: Hem URUNKODU hem de Düzenlenmiş Ürün Kodu ile
                                                urunkodu_codes = result_df['URUNKODU'].astype(str).tolist()
                                                duzenlenmis_codes = result_df['Düzenlenmiş Ürün Kodu'].astype(str).tolist()
                                                
                                                # Tam eşleşme kontrolü
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).apply(clean_product_code)
                                                duzenlenmis_clean = result_df['Düzenlenmiş Ürün Kodu'].astype(str).apply(clean_product_code)
                                                catalogue_clean = clean_product_code(catalogue_num)
                                                
                                                # Tam eşleşme
                                                match_mask_urun = urunkodu_clean == catalogue_clean
                                                match_mask_duzen = duzenlenmis_clean == catalogue_clean
                                                match_mask = match_mask_urun | match_mask_duzen
                                                
                                                # Eğer tam eşleşme yoksa, fuzzy matching dene
                                                if not match_mask.any():
                                                    best_match, best_ratio = find_best_match(catalogue_num, urunkodu_codes + duzenlenmis_codes, threshold=0.85)
                                                    if best_match and best_ratio >= 0.85:
                                                        # Fuzzy match bulundu, en yakın eşleşmeyi bul
                                                        fuzzy_match_mask = (urunkodu_clean == clean_product_code(best_match)) | (duzenlenmis_clean == clean_product_code(best_match))
                                                        if fuzzy_match_mask.any():
                                                            match_mask = fuzzy_match_mask
                                                            st.info(f"🔍 Schaeffler fuzzy match: {catalogue_num} → {best_match} (benzerlik: {best_ratio:.2f})")
                                
                                                # Debug: İlk 5 eşleştirme örneği göster
                                                if _ < 5:
                                                    match_count = match_mask.sum()
                                                    match_count_urun = match_mask_urun.sum()
                                                    match_count_duzen = match_mask_duzen.sum()
                                                    st.info(f"🔍 Schaeffler eşleştirme: {catalogue_num} → {match_count} eşleşme (URUNKODU: {match_count_urun}, Düzenlenmiş: {match_count_duzen})")
                                                    if match_count > 0:
                                                        # Eşleşen ürün kodlarını göster
                                                        matched_products = result_df.loc[match_mask, 'URUNKODU'].head(3).tolist()
                                                        st.write(f"    Eşleşen ürün kodları: {matched_products}")
                                
                                                if match_mask.sum() > 0:
                                                    # Tedarikçi kolonunu güncelle (toplama ile)
                                                    if tedarikci == 'İmes':
                                                        result_df.loc[match_mask, 'İmes Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaeffler {tedarikci}: {catalogue_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaeffler {tedarikci}: {catalogue_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaeffler {tedarikci}: {catalogue_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaeffler {tedarikci}: {catalogue_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'İkitelli':
                                                        result_df.loc[match_mask, 'İkitelli Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaeffler {tedarikci}: {catalogue_num} → {quantity} adet eklendi")
                                                else:
                                                    # Eşleşme bulunamadığında detaylı debug bilgisi
                                                    st.warning(f"⚠️ Schaeffler: {catalogue_num} için eşleşme bulunamadı")
                                                    st.write(f"  Temizlenmiş kod: {catalogue_clean}")
                                                    st.write(f"  Örnek URUNKODU: {urunkodu_clean.head(3).tolist()}")
                                                    st.write(f"  Örnek Düzenlenmiş: {duzenlenmis_clean.head(3).tolist()}")
                                
                                st.success(f"✅ Schaeffler Luk verileri işlendi: {len(schaeffler_df)} satır")
                            else:
                                st.warning("⚠️ Schaeffler dosyasında 'PO Number(L)' kolonu bulunamadı")
                                
                        except Exception as e:
                            st.error(f"❌ Schaeffler veri işleme hatası: {str(e)}")
                    
                    # ZF İthal için tedarikçi bakiye işlemi
                    elif 'ZF İTHAL' in brand:
                        try:
                            # Tedarikçi bakiye kolonlarını oluştur
                            if 'İmes Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İmes Tedarikçi Bakiye'] = 0
                            if 'Ankara Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarikçi Bakiye'] = 0
                            if 'Bolu Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarikçi Bakiye'] = 0
                            if 'Maslak Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarikçi Bakiye'] = 0
                            if 'İkitelli Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İkitelli Tedarikçi Bakiye'] = 0
                            
                            # ZF İthal verilerini işle
                            zf_ithal_df = brand_df.copy()
                            
                            # Material kolonunu kontrol et
                            if 'Material' in zf_ithal_df.columns:
                                # Material kodunu işle - düzeltilmiş kural
                                zf_ithal_df['Material_clean'] = zf_ithal_df['Material'].astype(str).apply(
                                    lambda x: x.split(':')[1].replace(' ', '') if ':' in x and (x.startswith('LF:') or x.startswith('SX:'))  # LF: veya SX: ile başlıyorsa : sonrasını al
                                    else x.split(':')[0].strip() if ':' in x and not (x.startswith('LF:') or x.startswith('SX:'))  # Diğerlerinde : öncesini al
                                    else x.replace(' ', '')  # : yoksa boşlukları sil
                                )
                                
                                # Debug: İlk 5 örnek göster
                                st.info(f"🔍 ZF İthal Material işleme örnekleri:")
                                for i, (original, cleaned) in enumerate(zip(zf_ithal_df['Material'].head(), zf_ithal_df['Material_clean'].head())):
                                    st.write(f"  {original} → {cleaned}")
                                
                                # Debug: Eşleştirme örnekleri
                                st.info(f"🔍 ZF İthal eşleştirme örnekleri:")
                                for i, material_num in enumerate(zf_ithal_df['Material_clean'].head()):
                                    urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip().str.replace(' ', '', regex=False).str.upper()
                                    duzenlenmis_clean = result_df['Düzenlenmiş Ürün Kodu'].astype(str).str.replace(' ', '', regex=False).str.upper()
                                    material_clean = material_num.replace(' ', '').upper()
                                    
                                    match_urun = (urunkodu_clean == material_clean).sum()
                                    match_duzen = (duzenlenmis_clean == material_clean).sum()
                                    
                                    st.write(f"  {material_num} → URUNKODU: {match_urun}, Düzenlenmiş: {match_duzen}")
                                
                                # Purchase order no. kolonunu kontrol et
                                if 'Purchase order no.' in zf_ithal_df.columns:
                                    # Tedarikçi kodlarını belirle
                                    zf_ithal_df['Tedarikçi'] = zf_ithal_df['Purchase order no.'].astype(str).apply(
                                        lambda x: 'İmes' if 'IME' in x or '285' in x or 'İST' in x or 'IST' in x
                                        else 'Ankara' if 'ANK' in x or '321' in x
                                        else 'Bolu' if '322' in x
                                        else 'Maslak' if '323' in x
                                        else 'İkitelli' if 'IKI' in x or '324' in x
                                        else 'Diğer'
                                    )
                                    
                                    # Debug: Tedarikçi dağılımını göster
                                    tedarikci_counts = zf_ithal_df['Tedarikçi'].value_counts()
                                    st.info(f"🔍 ZF İthal Tedarikçi dağılımı:")
                                    for tedarikci, count in tedarikci_counts.items():
                                        st.write(f"  {tedarikci}: {count} satır")
                                    
                                    # Debug: Örnek Purchase order no. kodları göster
                                    st.info("🔍 ZF İthal Purchase order no. örnekleri:")
                                    for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                        tedarikci_data = zf_ithal_df[zf_ithal_df['Tedarikçi'] == tedarikci]
                                        if len(tedarikci_data) > 0:
                                            sample_codes = tedarikci_data['Purchase order no.'].head(3).tolist()
                                            st.write(f"  {tedarikci}: {sample_codes}")
                                
                                # Qty.in Del. ve Open quantity kolonlarını kontrol et
                                if 'Qty.in Del.' in zf_ithal_df.columns and 'Open quantity' in zf_ithal_df.columns:
                                    # Tedarikçi bazında grupla ve topla
                                    for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                        tedarikci_data = zf_ithal_df[zf_ithal_df['Tedarikçi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Material_clean bazında topla
                                            grouped = tedarikci_data.groupby('Material_clean').agg({
                                                'Qty.in Del.': 'sum',
                                                'Open quantity': 'sum'
                                            }).reset_index()
                                            
                                            # Ana DataFrame ile eşleştir (LPR, Lemforder, TRW markaları)
                                            for _, row in grouped.iterrows():
                                                material_num = row['Material_clean']
                                                qty_del = row['Qty.in Del.']
                                                open_qty = row['Open quantity']
                                                total_qty = qty_del + open_qty
                                                
                                                # LEMFÖRDER, TRW, SACHS markalarını ara
                                                lemforder_mask = result_df['CAT4'].str.contains('LEMFÖRDER', case=False, na=False)
                                                trw_mask = result_df['CAT4'].str.contains('TRW', case=False, na=False)
                                                sachs_mask = result_df['CAT4'].str.contains('SACHS', case=False, na=False)
                                                
                                                # Hem URUNKODU hem de Düzenlenmiş Ürün Kodu ile tam eşleştir (case-insensitive)
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip().str.replace(' ', '', regex=False).str.upper()
                                                duzenlenmis_clean = result_df['Düzenlenmiş Ürün Kodu'].astype(str).str.replace(' ', '', regex=False).str.upper()
                                                material_clean = material_num.replace(' ', '').upper()
                                                
                                                # Tam eşleştirme yap (case-insensitive)
                                                match_mask_urun = urunkodu_clean == material_clean
                                                match_mask_duzen = duzenlenmis_clean == material_clean
                                                match_mask = match_mask_urun | match_mask_duzen
                                                
                                                # LEMFÖRDER, TRW, SACHS markaları ile birleştir
                                                final_mask = match_mask & (lemforder_mask | trw_mask | sachs_mask)
                                                
                                                if final_mask.sum() > 0:
                                                    # Tedarikçi kolonunu güncelle (toplama ile)
                                                    if tedarikci == 'İmes':
                                                        result_df.loc[final_mask, 'İmes Tedarikçi Bakiye'] += total_qty
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {total_qty} adet eklendi")
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[final_mask, 'Ankara Tedarikçi Bakiye'] += total_qty
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {total_qty} adet eklendi")
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[final_mask, 'Bolu Tedarikçi Bakiye'] += total_qty
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {total_qty} adet eklendi")
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[final_mask, 'Maslak Tedarikçi Bakiye'] += total_qty
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {total_qty} adet eklendi")
                                                    elif tedarikci == 'İkitelli':
                                                        result_df.loc[final_mask, 'İkitelli Tedarikçi Bakiye'] += total_qty
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {total_qty} adet eklendi")
                                                    
                                                    # Debug bilgisi
                                                    st.info(f"🔍 ZF İthal: {material_num} → {final_mask.sum()} eşleşme bulundu")
                                                else:
                                                    # Eşleşme bulunamadığında detaylı debug bilgisi
                                                    st.warning(f"⚠️ ZF İthal: {material_num} için eşleşme bulunamadı")
                                                    st.write(f"  Material (temiz): {material_num}")
                                                    st.write(f"  Material (boşluksuz): {material_no_space}")
                                                    
                                                    # Örnek URUNKODU ve Düzenlenmiş Ürün Kodu göster
                                                    sample_urun = result_df['URUNKODU'].head(5).tolist()
                                                    sample_duzen = result_df['Düzenlenmiş Ürün Kodu'].head(5).tolist()
                                                    st.write(f"  Örnek URUNKODU: {sample_urun}")
                                                    st.write(f"  Örnek Düzenlenmiş: {sample_duzen}")
                                
                                st.success(f"✅ ZF İthal verileri işlendi: {len(zf_ithal_df)} satır")
                            else:
                                st.warning("⚠️ ZF İthal dosyasında 'Material' kolonu bulunamadı")
                                
                        except Exception as e:
                            st.error(f"❌ ZF İthal veri işleme hatası: {str(e)}")
                
                    # ZF Yerli için tedarikçi bakiye işlemi
                    elif 'ZF YERLİ' in brand:
                        try:
                            # Tedarikçi bakiye kolonlarını oluştur
                            if 'İmes Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İmes Tedarikçi Bakiye'] = 0
                            if 'Ankara Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarikçi Bakiye'] = 0
                            if 'Bolu Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarikçi Bakiye'] = 0
                            if 'Maslak Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarikçi Bakiye'] = 0
                            if 'İkitelli Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İkitelli Tedarikçi Bakiye'] = 0
                            
                            # ZF Yerli verilerini işle
                            zf_yerli_df = brand_df.copy()
                            
                            # Basic No. kolonunu kontrol et
                            if 'Basic No.' in zf_yerli_df.columns:
                                # Basic No. kodunu temizle
                                zf_yerli_df['Basic_clean'] = zf_yerli_df['Basic No.'].astype(str).str.strip()
                                
                                # Ship-to Name kolonunu kontrol et
                                if 'Ship-to Name' in zf_yerli_df.columns:
                                    # Tedarikçi kodlarını belirle
                                    zf_yerli_df['Tedarikçi'] = zf_yerli_df['Ship-to Name'].astype(str).apply(
                                        lambda x: 'İmes' if 'IME' in x or '285' in x or 'IST' in x or 'İST' in x
                                        else 'Ankara' if 'ANK' in x or '321' in x
                                        else 'Bolu' if '322' in x
                                        else 'Maslak' if '323' in x
                                        else 'İkitelli' if 'IKI' in x or '324' in x
                                        else 'Diğer'
                                    )
                                    
                                    # Debug: Tedarikçi dağılımını göster
                                    tedarikci_counts = zf_yerli_df['Tedarikçi'].value_counts()
                                    st.info(f"🔍 ZF Yerli Tedarikçi dağılımı:")
                                    for tedarikci, count in tedarikci_counts.items():
                                        st.write(f"  {tedarikci}: {count} satır")
                                
                                # Outstanding Quantity kolonunu kontrol et
                                if 'Outstanding Quantity' in zf_yerli_df.columns:
                                    # Tedarikçi bazında grupla ve topla
                                    for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                        tedarikci_data = zf_yerli_df[zf_yerli_df['Tedarikçi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Basic_clean bazında topla
                                            grouped = tedarikci_data.groupby('Basic_clean')['Outstanding Quantity'].sum().reset_index()
                                            
                                            # Ana DataFrame ile eşleştir (Düzenlenmiş Ürün Kodu ile)
                                            for _, row in grouped.iterrows():
                                                basic_num = row['Basic_clean']
                                                quantity = row['Outstanding Quantity']
                                                
                                                # LEMFÖRDER, TRW, SACHS markalarını ara
                                                lemforder_mask = result_df['CAT4'].str.contains('LEMFÖRDER', case=False, na=False)
                                                trw_mask = result_df['CAT4'].str.contains('TRW', case=False, na=False)
                                                sachs_mask = result_df['CAT4'].str.contains('SACHS', case=False, na=False)
                                                
                                                # Düzenlenmiş Ürün Kodu ile tam eşleştir (case-insensitive, boşlukları temizle)
                                                duzenlenmis_clean = result_df['Düzenlenmiş Ürün Kodu'].astype(str).str.strip().str.replace(' ', '', regex=False).str.upper()
                                                basic_clean = basic_num.replace(' ', '').upper()
                                                match_mask = duzenlenmis_clean == basic_clean
                                                
                                                # LEMFÖRDER, TRW, SACHS markaları ile birleştir
                                                final_mask = match_mask & (lemforder_mask | trw_mask | sachs_mask)
                                                
                                                if final_mask.sum() > 0:
                                                    # Tedarikçi kolonunu güncelle (toplama ile)
                                                    if tedarikci == 'İmes':
                                                        result_df.loc[final_mask, 'İmes Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {basic_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[final_mask, 'Ankara Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {basic_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[final_mask, 'Bolu Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {basic_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[final_mask, 'Maslak Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {basic_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'İkitelli':
                                                        result_df.loc[final_mask, 'İkitelli Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {basic_num} → {quantity} adet eklendi")
                                
                                st.success(f"✅ ZF Yerli verileri işlendi: {len(zf_yerli_df)} satır")
                            else:
                                st.warning("⚠️ ZF Yerli dosyasında 'Basic No.' kolonu bulunamadı")
                                
                        except Exception as e:
                            st.error(f"❌ ZF Yerli veri işleme hatası: {str(e)}")
                
                    # Valeo için tedarikçi bakiye işlemi
                    elif 'VALEO' in brand:
                        try:
                            # Tedarikçi bakiye kolonlarını oluştur
                            if 'İmes Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İmes Tedarikçi Bakiye'] = 0
                            if 'Ankara Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarikçi Bakiye'] = 0
                            if 'Bolu Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarikçi Bakiye'] = 0
                            if 'Maslak Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarikçi Bakiye'] = 0
                            if 'İkitelli Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İkitelli Tedarikçi Bakiye'] = 0
                            
                            # Valeo verilerini işle
                            valeo_df = brand_df.copy()
                            
                            # Müşteri P/O No. kolonunu kontrol et
                            if 'Müşteri P/O No.' in valeo_df.columns:
                                # Tedarikçi kodlarını belirle
                                valeo_df['Tedarikçi'] = valeo_df['Müşteri P/O No.'].astype(str).apply(
                                    lambda x: 'İmes' if 'IME' in x or '285' in x
                                    else 'Ankara' if 'ANK' in x or '321' in x
                                    else 'Bolu' if '322' in x
                                    else 'Maslak' if '323' in x
                                    else 'İkitelli' if 'IKI' in x or '324' in x
                                    else 'Diğer'
                                )
                                
                                # Valeo Ref. kolonunu kontrol et - Geliştirilmiş
                                if 'Valeo Ref.' in valeo_df.columns:
                                    # Geliştirilmiş Valeo kod işleme
                                    valeo_df['Valeo_clean'] = valeo_df['Valeo Ref.'].apply(process_valeo_codes)
                                    
                                    # Debug: Valeo Ref. işleme örnekleri göster
                                    st.info(f"🔍 Valeo Ref. işleme örnekleri (Geliştirilmiş):")
                                    for i, (original, cleaned) in enumerate(zip(valeo_df['Valeo Ref.'].head(), valeo_df['Valeo_clean'].head())):
                                        st.write(f"  {original} → {cleaned}")
                                    
                                    # Eşleştirme istatistikleri
                                    total_codes = len(valeo_df['Valeo_clean'])
                                    unique_codes = valeo_df['Valeo_clean'].nunique()
                                    st.info(f"📊 Valeo kod istatistikleri: Toplam {total_codes}, Benzersiz {unique_codes}")
                                
                                # Sipariş Adeti kolonunu kontrol et
                                if 'Sipariş Adeti' in valeo_df.columns:
                                    # Tedarikçi bazında grupla ve topla
                                    for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                        tedarikci_data = valeo_df[valeo_df['Tedarikçi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Valeo_clean bazında topla
                                            grouped = tedarikci_data.groupby('Valeo_clean')['Sipariş Adeti'].sum().reset_index()
                                            
                                            # Ana DataFrame ile eşleştir - Geliştirilmiş
                                            for _, row in grouped.iterrows():
                                                valeo_ref = row['Valeo_clean']
                                                quantity = row['Sipariş Adeti']
                                                
                                                # Geliştirilmiş eşleştirme: Hem URUNKODU hem de Düzenlenmiş Ürün Kodu ile
                                                urunkodu_codes = result_df['URUNKODU'].astype(str).tolist()
                                                duzenlenmis_codes = result_df['Düzenlenmiş Ürün Kodu'].astype(str).tolist()
                                                
                                                # Tam eşleşme kontrolü
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).apply(clean_product_code)
                                                duzenlenmis_clean = result_df['Düzenlenmiş Ürün Kodu'].astype(str).apply(clean_product_code)
                                                valeo_clean = clean_product_code(valeo_ref)
                                                
                                                # Tam eşleşme
                                                match_mask_urun = urunkodu_clean == valeo_clean
                                                match_mask_duzen = duzenlenmis_clean == valeo_clean
                                                match_mask = match_mask_urun | match_mask_duzen
                                                
                                                # Eğer tam eşleşme yoksa, fuzzy matching dene
                                                if not match_mask.any():
                                                    best_match, best_ratio = find_best_match(valeo_ref, urunkodu_codes + duzenlenmis_codes, threshold=0.85)
                                                    if best_match and best_ratio >= 0.85:
                                                        # Fuzzy match bulundu, en yakın eşleşmeyi bul
                                                        fuzzy_match_mask = (urunkodu_clean == clean_product_code(best_match)) | (duzenlenmis_clean == clean_product_code(best_match))
                                                        if fuzzy_match_mask.any():
                                                            match_mask = fuzzy_match_mask
                                                            st.info(f"🔍 Valeo fuzzy match: {valeo_ref} → {best_match} (benzerlik: {best_ratio:.2f})")
                                
                                                # Debug: İlk 5 eşleştirme örneği göster
                                                if _ < 5:
                                                    match_count = match_mask.sum()
                                                    match_count_urun = match_mask_urun.sum()
                                                    match_count_duzen = match_mask_duzen.sum()
                                                    st.info(f"🔍 Valeo eşleştirme: {valeo_ref} → {match_count} eşleşme (URUNKODU: {match_count_urun}, Düzenlenmiş: {match_count_duzen})")
                                                    if match_count > 0:
                                                        # Eşleşen ürün kodlarını göster
                                                        matched_products = result_df.loc[match_mask, 'URUNKODU'].head(3).tolist()
                                                        st.write(f"    Eşleşen ürün kodları: {matched_products}")
                                
                                                if match_mask.sum() > 0:
                                                    # Tedarikçi kolonunu güncelle (toplama ile)
                                                    if tedarikci == 'İmes':
                                                        result_df.loc[match_mask, 'İmes Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Valeo {tedarikci}: {valeo_ref} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Valeo {tedarikci}: {valeo_ref} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Valeo {tedarikci}: {valeo_ref} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Valeo {tedarikci}: {valeo_ref} → {quantity} adet eklendi")
                                                    elif tedarikci == 'İkitelli':
                                                        result_df.loc[match_mask, 'İkitelli Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Valeo {tedarikci}: {valeo_ref} → {quantity} adet eklendi")
                                                else:
                                                    # Eşleşme bulunamadığında detaylı debug bilgisi
                                                    st.warning(f"⚠️ Valeo: {valeo_ref} için eşleşme bulunamadı")
                                                    st.write(f"  Temizlenmiş kod: {valeo_clean}")
                                                    st.write(f"  Örnek URUNKODU: {urunkodu_clean.head(3).tolist()}")
                                                    st.write(f"  Örnek Düzenlenmiş: {duzenlenmis_clean.head(3).tolist()}")
                                
                                st.success(f"✅ Valeo verileri işlendi: {len(valeo_df)} satır")
                            else:
                                st.warning("⚠️ Valeo dosyasında 'Müşteri P/O No.' kolonu bulunamadı")
                                
                        except Exception as e:
                            st.error(f"❌ Valeo veri işleme hatası: {str(e)}")
                
                    # Delphi için tedarikçi bakiye işlemi
                    elif 'DELPHI' in brand:
                        try:
                            # Tedarikçi bakiye kolonlarını oluştur
                            if 'İmes Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İmes Tedarikçi Bakiye'] = 0
                            if 'Ankara Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarikçi Bakiye'] = 0
                            if 'Bolu Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarikçi Bakiye'] = 0
                            if 'Maslak Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarikçi Bakiye'] = 0
                            if 'İkitelli Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İkitelli Tedarikçi Bakiye'] = 0
                            
                            # Delphi verilerini işle
                            delphi_df = brand_df.copy()
                            
                            # Şube kolonunu kontrol et
                            if 'Şube' in delphi_df.columns:
                                # Tedarikçi kodlarını belirle
                                delphi_df['Tedarikçi'] = delphi_df['Şube'].astype(str).apply(
                                    lambda x: 'Bolu' if 'Teknik Dizel-Bolu' in x
                                    else 'İmes' if 'Teknik Dizel-Ümraniye' in x
                                    else 'Maslak' if 'Teknik Dizel-Maslak' in x
                                    else 'Ankara' if 'Teknik Dizel-Ankara' in x
                                    else 'İkitelli' if 'Teknik Dizel-İkitelli' in x
                                    else 'Diğer'
                                )
                                
                                # Material kolonunu kontrol et
                                if 'Material' in delphi_df.columns:
                                    # Material kodunu temizle
                                    delphi_df['Material_clean'] = delphi_df['Material'].astype(str).str.strip()
                                    
                                    # Debug: Material kolonu işleme örnekleri göster
                                    st.info(f"🔍 Delphi Material kolonu işleme örnekleri:")
                                    for i, (original, cleaned) in enumerate(zip(delphi_df['Material'].head(), delphi_df['Material_clean'].head())):
                                        st.write(f"  {original} → {cleaned}")
                                
                                # Cum.qty kolonunu kontrol et
                                if 'Cum.qty' in delphi_df.columns:
                                    # Tedarikçi bazında grupla ve topla
                                    for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                        tedarikci_data = delphi_df[delphi_df['Tedarikçi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Material_clean bazında topla
                                            grouped = tedarikci_data.groupby('Material_clean')['Cum.qty'].sum().reset_index()
                                            
                                            # Ana DataFrame ile eşleştir
                                            for _, row in grouped.iterrows():
                                                material_num = row['Material_clean']
                                                quantity = row['Cum.qty']
                                                
                                                # Hem URUNKODU hem de Düzenlenmiş Ürün Kodu ile eşleştir
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip().str.replace(' ', '', regex=False).str.upper()
                                                duzenlenmis_clean = result_df['Düzenlenmiş Ürün Kodu'].astype(str).str.strip().str.replace(' ', '', regex=False).str.upper()
                                                material_clean = material_num.replace(' ', '').upper()
                                                
                                                # Her iki kolonla da eşleştir
                                                match_mask_urun = urunkodu_clean == material_clean
                                                match_mask_duzen = duzenlenmis_clean == material_clean
                                                match_mask = match_mask_urun | match_mask_duzen
                                                
                                                # Debug: İlk 5 eşleştirme örneği göster
                                                if _ < 5:
                                                    match_count = match_mask.sum()
                                                    match_count_urun = match_mask_urun.sum()
                                                    match_count_duzen = match_mask_duzen.sum()
                                                    st.info(f"🔍 Delphi eşleştirme: {material_num} → {match_count} eşleşme (URUNKODU: {match_count_urun}, Düzenlenmiş: {match_count_duzen})")
                                                
                                                if match_mask.sum() > 0:
                                                    # Tedarikçi kolonunu güncelle (toplama ile)
                                                    if tedarikci == 'İmes':
                                                        result_df.loc[match_mask, 'İmes Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Delphi {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Delphi {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Delphi {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Delphi {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'İkitelli':
                                                        result_df.loc[match_mask, 'İkitelli Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Delphi {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                else:
                                                    # Eşleşme bulunamadığında debug bilgisi
                                                    st.warning(f"⚠️ Delphi: {material_num} için eşleşme bulunamadı")
                                                    st.write(f"  Material (temiz): {material_clean}")
                                                    st.write(f"  Örnek URUNKODU: {urunkodu_clean.head(3).tolist()}")
                                                    st.write(f"  Örnek Düzenlenmiş: {duzenlenmis_clean.head(3).tolist()}")
                                
                                st.success(f"✅ Delphi verileri işlendi: {len(delphi_df)} satır")
                            else:
                                st.warning("⚠️ Delphi dosyasında 'Şube' kolonu bulunamadı")
                                
                        except Exception as e:
                            st.error(f"❌ Delphi veri işleme hatası: {str(e)}")
                
                    # Mann ve Filtron için tedarikçi bakiye işlemi
                    if 'MANN' in brand or 'FILTRON' in brand:
                        try:
                            # Tedarikçi bakiye kolonlarını oluştur
                            if 'İmes Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İmes Tedarikçi Bakiye'] = 0
                            if 'Ankara Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarikçi Bakiye'] = 0
                            if 'Bolu Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarikçi Bakiye'] = 0
                            if 'Maslak Tedarikçi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarikçi Bakiye'] = 0
                            if 'İkitelli Tedarikçi Bakiye' not in result_df.columns:
                                result_df['İkitelli Tedarikçi Bakiye'] = 0
                            
                            # Mann/Filtron verilerini işle
                            brand_df_processed = brand_df.copy()
                            
                            # Material Adı kolonunu kontrol et (farklı isimler için)
                            material_col = None
                            for col_name in ['Material Adı', 'Material', 'Material Name', 'Ürün Kodu', 'Product Code', 'Material Kodu', 'Malzeme Kodu', 'Malzeme Adı']:
                                if col_name in brand_df_processed.columns:
                                    material_col = col_name
                                    break
                            
                            if material_col:
                                st.info(f"🔍 {brand} için {material_col} kolonu bulundu!")
                                
                                # Debug: Tüm kolonları göster
                                st.info(f"🔍 {brand} dosyasındaki tüm kolonlar:")
                                for i, col in enumerate(brand_df_processed.columns):
                                    st.write(f"  {i+1}. {col}")
                                
                                # Debug: Önemli kolonları kontrol et
                                important_cols = ['Müşteri SatınAlma No', 'Açık Sipariş Adedi', 'Material Kodu', 'Material Adı']
                                st.info(f"🔍 {brand} önemli kolonlar kontrolü:")
                                for col in important_cols:
                                    if col in brand_df_processed.columns:
                                        st.success(f"  ✅ {col} - BULUNDU")
                                    else:
                                        st.error(f"  ❌ {col} - BULUNAMADI")
                                
                                # Material kodunu temizle (bulunan kolon adını kullan)
                                brand_df_processed['Material_clean'] = brand_df_processed[material_col].astype(str).str.strip()
                                
                                # Debug: Material kolonu örnekleri göster
                                st.info(f"🔍 {brand} Material kolonu örnekleri:")
                                sample_materials = brand_df_processed[material_col].head(10).tolist()
                                for i, material in enumerate(sample_materials):
                                    st.write(f"  {i+1}. {material}")
                                
                                # Debug: Material_clean örnekleri göster
                                st.info(f"🔍 {brand} Material_clean örnekleri:")
                                sample_cleaned = brand_df_processed['Material_clean'].head(10).tolist()
                                for i, cleaned in enumerate(sample_cleaned):
                                    st.write(f"  {i+1}. {cleaned}")
                                
                                # Müşteri SatınAlma No kolonunu kontrol et
                                if 'Müşteri SatınAlma No' in brand_df_processed.columns:
                                    # Debug: Müşteri SatınAlma No örnekleri göster
                                    st.info(f"🔍 {brand} Müşteri SatınAlma No örnekleri:")
                                    sample_codes = brand_df_processed['Müşteri SatınAlma No'].head(10).tolist()
                                    for i, code in enumerate(sample_codes):
                                        st.write(f"  {i+1}. {code}")
                                    
                                    # Tedarikçi kodlarını belirle
                                    brand_df_processed['Tedarikçi'] = brand_df_processed['Müşteri SatınAlma No'].astype(str).apply(
                                        lambda x: 'Ankara' if 'AAS' in x
                                        else 'İmes' if 'DAS' in x
                                        else 'Bolu' if 'BAS' in x
                                        else 'Maslak' if 'MAS' in x
                                        else 'İkitelli' if 'EAS' in x
                                        else 'Diğer'
                                    )
                                    
                                    # Debug: Tedarikçi dağılımı göster
                                    st.info(f"🔍 {brand} tedarikçi dağılımı:")
                                    tedarikci_dist = brand_df_processed['Tedarikçi'].value_counts()
                                    for tedarikci, count in tedarikci_dist.items():
                                        st.write(f"  {tedarikci}: {count} adet")
                                
                                # Açık Sipariş Adedi kolonunu kontrol et
                                if 'Açık Sipariş Adedi' in brand_df_processed.columns:
                                    # Tedarikçi bazında grupla ve topla
                                    for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                        tedarikci_data = brand_df_processed[brand_df_processed['Tedarikçi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Material_clean bazında topla
                                            grouped = tedarikci_data.groupby('Material_clean')['Açık Sipariş Adedi'].sum().reset_index()
                                            
                                            # Ana DataFrame ile eşleştir
                                            for _, row in grouped.iterrows():
                                                material_num = row['Material_clean']
                                                quantity = row['Açık Sipariş Adedi']
                                                
                                                # Hem URUNKODU hem de Düzenlenmiş Ürün Kodu ile tam eşleştir (case-insensitive)
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip().str.replace(' ', '', regex=False).str.upper()
                                                duzenlenmis_clean = result_df['Düzenlenmiş Ürün Kodu'].astype(str).str.replace(' ', '', regex=False).str.upper()
                                                material_clean_no_space = material_num.replace(' ', '').upper()
                                                
                                                # Tam eşleştirme yap (case-insensitive)
                                                match_mask_urun = urunkodu_clean == material_clean_no_space
                                                match_mask_duzen = duzenlenmis_clean == material_clean_no_space
                                                match_mask = match_mask_urun | match_mask_duzen
                                                
                                                # Debug: Eşleştirme detayları
                                                st.info(f"🔍 {brand} tam eşleştirme (case-insensitive): {material_num} → {material_clean_no_space}")
                                                st.info(f"  URUNKODU tam eşleşme: {match_mask_urun.sum()} adet")
                                                st.info(f"  Düzenlenmiş Ürün Kodu tam eşleşme: {match_mask_duzen.sum()} adet")
                                                st.info(f"  Toplam tam eşleşme: {match_mask.sum()} adet")
                                                
                                                if match_mask.sum() > 0:
                                                    # Tedarikçi kolonunu güncelle (toplama ile)
                                                    if tedarikci == 'İmes':
                                                        result_df.loc[match_mask, 'İmes Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                    elif tedarikci == 'İkitelli':
                                                        result_df.loc[match_mask, 'İkitelli Tedarikçi Bakiye'] += quantity
                                                        st.success(f"✅ Schaflerr {tedarikci}: {material_num} → {quantity} adet eklendi")
                                                else:
                                                    # Eşleşme bulunamadığında debug bilgisi
                                                    st.warning(f"⚠️ Schaflerr: {material_num} için eşleşme bulunamadı")
                                
                                st.success(f"✅ {brand} verileri işlendi: {len(brand_df_processed)} satır")
                                
                                # Debug: Sonuç kontrolü
                                st.info(f"🔍 {brand} sonuç kontrolü:")
                                for tedarikci in ['İmes', 'Ankara', 'Bolu', 'Maslak', 'İkitelli']:
                                    col_name = f"{tedarikci} Tedarikçi Bakiye"
                                    if col_name in result_df.columns:
                                        total = result_df[col_name].sum()
                                        st.write(f"  {tedarikci}: {total} adet")
                            else:
                                st.warning(f"⚠️ {brand} dosyasında gerekli kolonlar bulunamadı. Mevcut kolonlar: {list(brand_df_processed.columns)}")
                                
                        except Exception as e:
                            st.error(f"❌ {brand} veri işleme hatası: {str(e)}")
                
                if brand_count == 0:
                    st.warning(f"⚠️ {brand} markası CAT4 kolonunda bulunamadı")
        
        # Marka eşleştirme sonrası toplam depo bakiyesi güncelleme
        depo_bakiye_cols = ['Maslak Depo Bakiye', 'Bolu Depo Bakiye', 'İmes Depo Bakiye', 'Ankara Depo Bakiye', 'İkitelli Depo Bakiye']
        available_depo_cols = [col for col in depo_bakiye_cols if col in result_df.columns]
        
        if available_depo_cols and 'Toplam Depo Bakiye' in result_df.columns:
            # Sayısal değerlere çevir ve topla
            for col in available_depo_cols:
                result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0)
            
            # Toplam hesapla
            result_df['Toplam Depo Bakiye'] = result_df[available_depo_cols].sum(axis=1)
            
            st.success(f"✅ Toplam Depo Bakiye hesaplandı: {len(available_depo_cols)} depo kolonu toplandı")
        
        # Tedarikçi bakiye toplamlarını göster
        tedarikci_cols = ['İmes Tedarikçi Bakiye', 'Ankara Tedarikçi Bakiye', 'Bolu Tedarikçi Bakiye', 'Maslak Tedarikçi Bakiye', 'İkitelli Tedarikçi Bakiye']
        available_tedarikci_cols = [col for col in tedarikci_cols if col in result_df.columns]
        
        if available_tedarikci_cols:
            st.info("🔍 Tedarikçi Bakiye Toplamları:")
            for col in available_tedarikci_cols:
                total = result_df[col].sum()
                st.write(f"  {col}: {total:,.0f} adet")
        
        return result_df
        
    except Exception as e:
        st.error(f"Marka eşleştirme hatası: {str(e)}")
        return main_df

@st.cache_data(show_spinner="Excel oluşturuluyor...", ttl=1800)
def format_excel_ultra_fast(df):
    """Ultra hızlı Excel oluşturma - performans odaklı"""
    try:
        output = BytesIO()
        
        # DataFrame'i kopyala ve "-" değerlerini 0'a çevir
        df_clean = df.copy()
        
        # Depo ve tedarikçi bakiye kolonlarında "-" değerlerini 0'a çevir
        depo_cols = [col for col in df_clean.columns if any(keyword in col for keyword in 
                   ['DEVIR', 'ALIŞ', 'SATIS', 'STOK', 'Depo Bakiye', 'Tedarikçi Bakiye'])]
        
        for col in depo_cols:
            if col in df_clean.columns:
                # Önce string'e çevir, sonra temizlik yap
                df_clean[col] = df_clean[col].astype(str)
                df_clean[col] = df_clean[col].replace('-', '0')
                df_clean[col] = df_clean[col].replace('nan', '0')
                df_clean[col] = df_clean[col].replace('None', '0')
                
                # Sayısal değerlere çevir
                df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0)
        
        # Debug: Temizlenen kolonları göster
        st.info(f"🔧 Temizlenen kolonlar: {len(depo_cols)} adet")
        for col in depo_cols[:5]:  # İlk 5 kolonu göster
            st.write(f"  - {col}")
        if len(depo_cols) > 5:
            st.write(f"  ... ve {len(depo_cols)-5} kolon daha")
        
        # Her zaman performans modu kullan - hız için
        # Excel oluşturma ve özel format uygulama
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_clean.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Excel workbook ve worksheet'i al
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Düzenlenmiş Ürün Kodu kolonuna özel format uygula
            for col_num, col_name in enumerate(df_clean.columns, 1):
                if col_name == 'Düzenlenmiş Ürün Kodu':
                    # Bu kolon için özel format: metin formatı
                    for row_num in range(2, len(df_clean) + 2):  # Excel'de satır 1 başlık
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.number_format = '@'  # Metin formatı
                    break
            
            # Toplam Depo Bakiye kolonuna formül ekle
            toplam_depo_col = None
            depo_bakiye_cols = []
            
            # Depo bakiye kolonlarını bul
            for col_num, col_name in enumerate(df_clean.columns, 1):
                if 'Depo Bakiye' in col_name and col_name != 'Toplam Depo Bakiye':
                    depo_bakiye_cols.append(col_name)
                elif col_name == 'Toplam Depo Bakiye':
                    toplam_depo_col = col_num
            
            # Formül ekle
            if toplam_depo_col and depo_bakiye_cols:
                for row_num in range(2, len(df_clean) + 2):  # Excel'de satır 1 başlık
                    cell = worksheet.cell(row=row_num, column=toplam_depo_col)
                    
                    # Formül oluştur: =SUM(Maslak Depo Bakiye:Bolu Depo Bakiye:İmes Depo Bakiye:Ankara Depo Bakiye:İkitelli Depo Bakiye)
                    formula_parts = []
                    for depo_col in depo_bakiye_cols:
                        # Kolon harfini bul
                        for col_idx, col_name in enumerate(df_clean.columns, 1):
                            if col_name == depo_col:
                                col_letter = chr(64 + col_idx)  # A=65, B=66, etc.
                                if col_idx > 26:
                                    col_letter = chr(64 + (col_idx // 26)) + chr(64 + (col_idx % 26))
                                formula_parts.append(f"{col_letter}{row_num}")
                                break
                    
                    if formula_parts:
                        formula = f"=SUM({','.join(formula_parts)})"
                        cell.value = formula
        
        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        # Hata durumunda da Excel oluştur
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Excel workbook ve worksheet'i al
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Düzenlenmiş Ürün Kodu kolonuna özel format uygula
            for col_num, col_name in enumerate(df.columns, 1):
                if col_name == 'Düzenlenmiş Ürün Kodu':
                    # Bu kolon için özel format: metin formatı
                    for row_num in range(2, len(df) + 2):  # Excel'de satır 1 başlık
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.number_format = '@'  # Metin formatı
                    break
            
            # Toplam Depo Bakiye kolonuna formül ekle (hata durumunda)
            toplam_depo_col = None
            depo_bakiye_cols = []
            
            # Depo bakiye kolonlarını bul
            for col_num, col_name in enumerate(df.columns, 1):
                if 'Depo Bakiye' in col_name and col_name != 'Toplam Depo Bakiye':
                    depo_bakiye_cols.append(col_name)
                elif col_name == 'Toplam Depo Bakiye':
                    toplam_depo_col = col_num
            
            # Formül ekle
            if toplam_depo_col and depo_bakiye_cols:
                for row_num in range(2, len(df) + 2):  # Excel'de satır 1 başlık
                    cell = worksheet.cell(row=row_num, column=toplam_depo_col)
                    
                    # Formül oluştur
                    formula_parts = []
                    for depo_col in depo_bakiye_cols:
                        # Kolon harfini bul
                        for col_idx, col_name in enumerate(df.columns, 1):
                            if col_name == depo_col:
                                col_letter = chr(64 + col_idx)  # A=65, B=66, etc.
                                if col_idx > 26:
                                    col_letter = chr(64 + (col_idx // 26)) + chr(64 + (col_idx % 26))
                                formula_parts.append(f"{col_letter}{row_num}")
                                break
                    
                    if formula_parts:
                        formula = f"=SUM({','.join(formula_parts)})"
                        cell.value = formula
        
        output.seek(0)
        return output.getvalue()

# Ana uygulama
def main():
    # Hata yakalama ve yeniden başlatma kontrolü
    if 'kerim_restarted' not in st.session_state:
        st.session_state.kerim_restarted = False
    
    # Eğer sayfa yeniden başlatıldıysa
    if st.session_state.kerim_restarted:
        st.success("✅ Sayfa başarıyla yeniden başlatıldı!")
        st.session_state.kerim_restarted = False
    
    # Dosya yükleme alanı
    with st.expander("📤 ANA EXCEL DOSYASINI YÜKLEYİN", expanded=True):
        uploaded_file = st.file_uploader(
            "Excel dosyasını seçin (XLSX/XLS)",
            type=['xlsx', 'xls'],
            key="main_file"
        )
    
    if uploaded_file:
        try:
            # Hızlı işlem akışı
            with st.spinner("⚡ Dosya işleniyor..."):
                # 1. Hızlı okuma
                df = load_data_ultra_fast(uploaded_file)
                st.success(f"✅ Yüklendi: {len(df):,} satır | {len(df.columns)} sütun")
                
                # 2. Hızlı dönüşüm
                transformed_df = transform_data_ultra_fast(df)
                st.session_state.processed_data = transformed_df
                
                # 3. Hızlı Excel oluşturma
                if transformed_df is not None and len(transformed_df) > 0:
                    try:
                        excel_data = format_excel_ultra_fast(transformed_df)
                        st.download_button(
                            label=f"📥 Dönüştürülmüş Veriyi İndir ({len(transformed_df):,} satır)",
                            data=excel_data,
                            file_name=f"donusturulmus_veri_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                    except Exception as e:
                        st.error(f"Excel oluşturma hatası: {str(e)}")
                else:
                    st.warning("Dönüştürülecek veri bulunamadı.")
        
        except Exception as e:
            st.error(f"❌ Hata: {str(e)}")
            st.error("💡 Çözüm: Cache temizleyin veya sayfayı yenileyin.")
            
            # Cache temizleme ve yeniden başlatma butonları
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("🧹 Cache Temizle", type="secondary"):
                    if clear_all_caches():
                        st.success("✅ Cache temizlendi!")
                        st.rerun()
                    else:
                        st.error("❌ Cache temizleme başarısız!")
            
            with col2:
                if st.button("🔄 Sayfayı Yeniden Başlat", type="secondary"):
                    st.session_state.kerim_restarted = True
                    st.rerun()
            
            st.stop()
    
    # 7 farklı Excel ekleme kutusu - hızlı yükleme
    st.header("📂 Ek Excel Dosyalarını Yükleme")
    st.write("Aşağıdaki 7 Excel dosyasını yükleyin:")
    
    # 7 Excel dosyası yükleme - tek sütun
    excel1 = st.file_uploader("Schaeffler Luk", type=['xlsx', 'xls'], key="excel1")
    excel2 = st.file_uploader("ZF İthal Bakiye", type=['xlsx', 'xls'], key="excel2")
    excel3 = st.file_uploader("Delphi Bakiye", type=['xlsx', 'xls'], key="excel3")
    excel4 = st.file_uploader("ZF Yerli Bakiye", type=['xlsx', 'xls'], key="excel4")
    excel5 = st.file_uploader("Valeo Bakiye", type=['xlsx', 'xls'], key="excel5")
    excel6 = st.file_uploader("Filtron Bakiye", type=['xlsx', 'xls'], key="excel6")
    excel7 = st.file_uploader("Mann Bakiye", type=['xlsx', 'xls'], key="excel7")
    
    # Yükleme kontrolü
    uploaded_files = {
        'excel1': excel1, 'excel2': excel2, 'excel3': excel3, 'excel4': excel4,
        'excel5': excel5, 'excel6': excel6, 'excel7': excel7
    }
    uploaded_count = sum(1 for file in uploaded_files.values() if file is not None)
    
    st.write(f"**Yüklenen dosya sayısı:** {uploaded_count}/7")
    
    # Güncelle butonu
    if uploaded_count > 0:
        if st.button("🚀 Ultra Hızlı Marka Eşleştirme Yap", type="primary"):
            try:
                if st.session_state.processed_data is not None:
                    # Paralel marka eşleştirme işlemi
                    with st.spinner("⚡ Marka eşleştirme yapılıyor..."):
                        final_df = match_brands_parallel(st.session_state.processed_data, uploaded_files)
                        st.success(f"✅ Ultra hızlı marka eşleştirme tamamlandı! {len(final_df)} satır işlendi.")
                    
                    # Final Excel indirme butonu
                    if len(final_df) > 0:
                        try:
                            with st.spinner("⚡ Final Excel oluşturuluyor..."):
                                final_excel_data = format_excel_ultra_fast(final_df)
                                st.download_button(
                                    label=f"📥 Eşleştirilmiş Veriyi İndir ({len(final_df):,} satır)",
                                    data=final_excel_data,
                                    file_name=f"eslestirilmis_veri_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )
                        except Exception as e:
                            st.error(f"Final Excel oluşturma hatası: {str(e)}")
                            st.error("💡 Çözüm: Sayfayı yenileyin ve tekrar deneyin.")
                else:
                    st.warning("Önce ana Excel dosyasını yükleyin ve dönüştürün.")
            except Exception as e:
                st.error(f"❌ Marka eşleştirme hatası: {str(e)}")
                st.error("💡 Çözüm: Cache temizleyin veya sayfayı yenileyin.")
                
                # Cache temizleme ve yeniden başlatma butonları
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("🧹 Cache Temizle", type="secondary"):
                        if clear_all_caches():
                            st.success("✅ Cache temizlendi!")
                            st.rerun()
                        else:
                            st.error("❌ Cache temizleme başarısız!")
                
                with col2:
                    if st.button("🔄 Sayfayı Yeniden Başlat", type="secondary"):
                        st.session_state.kerim_restarted = True
                        st.rerun()
    else:
        st.info("Lütfen en az bir marka dosyası yükleyin.")
    
    # Ana sayfaya dönüş ve cache temizleme
    st.markdown("---")
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🏠 Ana Sayfaya Dön", type="secondary"):
            st.switch_page("Home")
    
    with col2:
        if st.button("🧹 Cache Temizle", type="secondary"):
            if clear_all_caches():
                st.success("✅ Cache başarıyla temizlendi!")
                st.session_state.app_restart_count += 1
                st.rerun()
            else:
                st.error("❌ Cache temizleme başarısız!")

# Sidebar
def sidebar():
    st.sidebar.header("⚡ Maksimum Hız Modu")
    
    st.sidebar.success("""
    **Aktif Optimizasyonlar:**
    - Minimal dtype belirtme
    - Vektörel işlemler
    - Bellek optimizasyonu
    - Paralel işleme
    - Hızlı Excel oluşturma
    """)
    
    # Cache durumu
    st.sidebar.markdown("---")
    st.sidebar.subheader("🧹 Cache Durumu")
    
    if st.sidebar.button("Cache Temizle", type="secondary"):
        if clear_all_caches():
            st.sidebar.success("✅ Cache temizlendi!")
            st.rerun()
        else:
            st.sidebar.error("❌ Cache temizleme başarısız!")
    
    # Yeniden başlatma sayısı
    restart_count = st.session_state.get('app_restart_count', 0)
    st.sidebar.info(f"🔄 Yeniden başlatma sayısı: {restart_count}")
    
    st.sidebar.header("📋 Kurallar")
    st.sidebar.write("- 0 değerleri → '-' olarak değiştirilir")
    st.sidebar.write("- Depo önekleri yeni isimlere dönüştürülür")
    st.sidebar.write("- Kategori sütunları korunur")
    st.sidebar.write("- Vektörel işlemler ile hızlandırma")
    st.sidebar.write("- Geliştirilmiş ürün kodu eşleştirme")
    st.sidebar.write("- Fuzzy matching desteği")
    
    st.sidebar.header("ℹ️ Bilgi")
    st.sidebar.write(f"Son Güncelleme: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}")

if __name__ == "__main__":
    sidebar()
    main() 