import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from functools import lru_cache

# Cache temizleme fonksiyonu
def clear_all_caches():
    """T√ºm cache'leri temizle"""
    try:
        # Cache temizleme
        st.cache_data.clear()
        st.cache_resource.clear()
        
        # Session state temizleme
        if 'processed_data' in st.session_state:
            del st.session_state.processed_data
        if 'brand_data_cache' in st.session_state:
            del st.session_state.brand_data_cache
        
        return True
    except Exception as e:
        st.error(f"Cache temizleme hatasƒ±: {str(e)}")
        return False

# Sayfa ayarlarƒ±
st.set_page_config(
    page_title="Excel D√∂n√º≈üt√ºrme Aracƒ± (Ultra Hƒ±zlƒ±)",
    page_icon="‚ö°",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Ba≈ülƒ±k
st.title("‚ö° Ultra Hƒ±zlƒ± Excel D√∂n√º≈üt√ºr√ºc√º")
st.caption("100.000+ satƒ±rlƒ±k dosyalar i√ßin optimize edilmi≈ü versiyon - Maksimum Hƒ±z Modu")

# Global deƒüi≈ükenler
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'brand_data_cache' not in st.session_state:
    st.session_state.brand_data_cache = {}
if 'app_restart_count' not in st.session_state:
    st.session_state.app_restart_count = 0

# Ultra hƒ±zlƒ± √∂nbellek fonksiyonlarƒ±
@st.cache_data(max_entries=5, show_spinner="Dosya okunuyor...", ttl=3600)
def load_data_ultra_fast(uploaded_file):
    """Maksimum hƒ±zlƒ± dosya okuma"""
    try:
        # Maksimum hƒ±z i√ßin minimal ayarlar
        df = pd.read_excel(
            uploaded_file,
            engine='openpyxl',
            # dtype belirtme - sadece kritik s√ºtunlar
            dtype={
                'URUNKODU': 'string'
            },
            # NaN kontrol√º tamamen devre dƒ±≈üƒ±
            na_filter=False,
            keep_default_na=False,
            # Ek hƒ±zlandƒ±rma
            header=0,
            skiprows=None,
            nrows=None  # T√ºm satƒ±rlarƒ± oku
        )
        
        return df
    except Exception as e:
        st.error(f"Dosya okuma hatasƒ±: {str(e)}")
        return pd.DataFrame()

@st.cache_data(show_spinner="Marka verisi okunuyor...", ttl=1800)
def load_brand_data_parallel(excel_file, brand_name):
    """Maksimum hƒ±zlƒ± marka verisi okuma"""
    try:
        # Maksimum hƒ±z i√ßin minimal ayarlar
        df = pd.read_excel(
            excel_file,
            engine='openpyxl',
            na_filter=False,
            keep_default_na=False
        )
        
        return brand_name, df
    except Exception as e:
        return brand_name, pd.DataFrame()

@st.cache_data(show_spinner="Veri d√∂n√º≈üt√ºr√ºl√ºyor...", ttl=3600)
def transform_data_ultra_fast(df):
    """Maksimum hƒ±zlƒ± veri d√∂n√º≈üt√ºrme"""
    try:
        # Sadece gerekli s√ºtunlarƒ± al - bellek tasarrufu
        essential_cols = [
            'URUNKODU', 'ACIKLAMA', 'URETƒ∞Cƒ∞KODU', 'ORJƒ∞NAL', 'ESKƒ∞KOD',
            'TOPL.FAT.ADT', 'M√ú≈ûT.SAY.', 'SATƒ±≈û FIYATƒ±', 'D√ñVIZ CINSI (S)'
        ] + [f'CAT{i}' for i in range(1, 8)]
        
        # Depo s√ºtunlarƒ± - sadece mevcut olanlarƒ± al
        depo_prefixes = ['02-', '04-', 'D01-', 'A01-', 'TD-E01-', 'E01-']
        depo_cols = []
        for prefix in depo_prefixes:
            for col_type in ['DEVIR', 'ALIS', 'STOK', 'SATIS']:
                col_name = f"{prefix}{col_type}"
                if col_name in df.columns:
                    depo_cols.append(col_name)
        
        # Mevcut s√ºtunlarƒ± filtrele
        available_cols = [col for col in essential_cols + depo_cols if col in df.columns]
        df_filtered = df[available_cols].copy()
        
        # Maksimum hƒ±zlƒ± d√∂n√º≈ü√ºm - vekt√∂rel i≈ülemler
        new_df = pd.DataFrame()
        
        # 1. URUNKODU (ilk) - vekt√∂rel
        new_df['URUNKODU'] = df_filtered['URUNKODU'].fillna(0)
        
        # 2. D√ºzenlenmi≈ü √úr√ºn Kodu - vekt√∂rel (ba≈üƒ±nda 0 olan kodlar i√ßin √∂zel format)
        new_df['D√ºzenlenmi≈ü √úr√ºn Kodu'] = df_filtered['URUNKODU'].fillna(0).str.replace(r'^[^-]*-', "", regex=True)
        
        # 4-7. Temel s√ºtunlar - vekt√∂rel
        basic_cols = ['ACIKLAMA', 'URETƒ∞Cƒ∞KODU', 'ORJƒ∞NAL', 'ESKƒ∞KOD']
        for col in basic_cols:
            if col in df_filtered.columns:
                new_df[col] = df_filtered[col].fillna(0)
        
        # 8. Kategoriler - vekt√∂rel
        for i in range(1, 8):
            cat_col = f'CAT{i}'
            if cat_col in df_filtered.columns:
                new_df[f'CAT{i}'] = df_filtered[cat_col].fillna(0)
        
        # 9. Depo verileri - vekt√∂rel i≈ülem
        depo_mapping = {
            '02-': 'MASLAK',
            'D01-': 'ƒ∞MES',
            'TD-E01-': 'ƒ∞Kƒ∞TELLƒ∞',
            'E01-': 'ƒ∞Kƒ∞TELLƒ∞',
            '04-': 'BOLU',
            'A01-': 'ANKARA'
        }
        
        for old_prefix, new_name in depo_mapping.items():
            for col_type, new_type in zip(['DEVIR', 'ALIS', 'SATIS', 'STOK'],
                                         ['DEVIR', 'ALI≈û', 'SATIS', 'STOK']):
                old_col = f"{old_prefix}{col_type}"
                if old_col in df_filtered.columns:
                    # Vekt√∂rel i≈ülem
                    col_data = df_filtered[old_col].fillna(0)
                    if pd.api.types.is_numeric_dtype(col_data):
                        col_data = col_data.astype(float)
                        col_data = col_data.replace(0, '-')
                    else:
                        col_data = col_data.astype(str)
                    new_df[f"{new_name} {new_type}"] = col_data.astype('string')
                else:
                    # Eksik s√ºtun i√ßin bo≈ü deƒüer
                    new_df[f"{new_name} {new_type}"] = '-'
        
        # 10. Tedarik√ßi bakiye kolonlarƒ± - vekt√∂rel
        tedarikci_cols = [
            'ƒ∞mes Tedarik√ßi Bakiye', 'Ankara Tedarik√ßi Bakiye', 
            'Bolu Tedarik√ßi Bakiye', 'Maslak Tedarik√ßi Bakiye'
        ]
        
        for col in tedarikci_cols:
            new_df[col] = '-'
        
        # 11. Dinamik ay ba≈ülƒ±klarƒ± - vekt√∂rel
        current_month = datetime.datetime.now().month
        months = ['Ocak', '≈ûubat', 'Mart', 'Nisan', 'Mayƒ±s', 'Haziran',
                 'Temmuz', 'Aƒüustos', 'Eyl√ºl', 'Ekim', 'Kasƒ±m', 'Aralƒ±k']
        
        next_month1 = months[(current_month) % 12]
        next_month2 = months[(current_month + 1) % 12]
        
        # Vekt√∂rel ay ba≈ülƒ±klarƒ±
        for i in range(5):
            new_df[f'{next_month1}_{i+1}'] = 0
            new_df[f'{next_month2}_{i+1}'] = 0
        
        # 12. Diƒüer s√ºtunlar - vekt√∂rel
        other_cols = {
            'TOPL.FAT.ADT': 'TOPL.FAT.ADT',
            'M√ú≈ûT.SAY.': 'M√ú≈ûT.SAY.',
            'SATƒ±≈û FIYATƒ±': 'SATƒ±≈û FIYATƒ±',
            'D√ñVIZ CINSI (S)': 'D√ñVIZ CINSI (S)'
        }
        
        for old, new in other_cols.items():
            if old in df_filtered.columns:
                new_df[new] = df_filtered[old].fillna(0)
        
        # 13. URUNKODU (D√ñVIZ CINSI'den sonra)
        new_df['URUNKODU_3'] = df_filtered['URUNKODU'].fillna(0)
        
        # 14. Eksik ba≈ülƒ±klarƒ± geri getir - vekt√∂rel
        # not, ƒ∞SK, PRƒ∞M, B√úT√áE, liste, TD SF, Net Fiyat Kampanyasƒ±
        new_df['not'] = 0
        new_df['ƒ∞SK'] = 0
        new_df['PRƒ∞M'] = 0
        new_df['B√úT√áE'] = 0
        new_df['liste'] = 0
        new_df['TD SF'] = 0
        new_df['Net Fiyat Kampanyasƒ±'] = 0
        
        # Kampanya Tipi
        new_df['Kampanya Tipi'] = 0
        
        # Toplam ƒ∞sk
        new_df['Toplam ƒ∞sk'] = 0
        
        # Depo Bakiye kolonlarƒ±
        new_df['Maslak Depo Bakiye'] = 0
        new_df['Bolu Depo Bakiye'] = 0
        new_df['ƒ∞mes Depo Bakiye'] = 0
        new_df['Ankara Depo Bakiye'] = 0
        new_df['ƒ∞kitelli Depo Bakiye'] = 0
        
        # Toplam Depo Bakiye - otomatik hesaplama
        new_df['Toplam Depo Bakiye'] = 0
        
        # Tedarik√ßi bakiye kolonlarƒ± - ƒ∞kitelli Tedarik√ßi Bakiye eklendi
        tedarikci_cols = [
            'ƒ∞mes Tedarik√ßi Bakiye', 'Ankara Tedarik√ßi Bakiye', 
            'Bolu Tedarik√ßi Bakiye', 'Maslak Tedarik√ßi Bakiye', 'ƒ∞kitelli Tedarik√ßi Bakiye'
        ]
        
        for col in tedarikci_cols:
            new_df[col] = 0
        
        # Paket Adetleri
        new_df['Paket Adetleri'] = 0
        
        # Sipari≈ü kolonlarƒ±
        new_df['Maslak Sipari≈ü'] = 0
        new_df['Bolu Sipari≈ü'] = 0
        new_df['ƒ∞mes Sipari≈ü'] = 0
        new_df['Ankara Sipari≈ü'] = 0
        new_df['ƒ∞kitelli Sipari≈ü'] = 0
        
        # S√ºtun sƒ±ralamasƒ±nƒ± d√ºzelt - verilen sƒ±raya g√∂re (64 adet)
        desired_order = [
            'URUNKODU', 'D√ºzenlenmi≈ü √úr√ºn Kodu', 'ACIKLAMA', 'URETƒ∞Cƒ∞KODU', 'ORJƒ∞NAL', 'ESKƒ∞KOD',
            'CAT1', 'CAT2', 'CAT3', 'CAT4', 'CAT5', 'CAT6', 'CAT7',
            # Depo kolonlarƒ± (sƒ±ralama: MASLAK, ƒ∞MES, ƒ∞Kƒ∞TELLƒ∞, BOLU, ANKARA)
            'MASLAK DEVIR', 'MASLAK ALI≈û', 'MASLAK SATIS', 'MASLAK STOK',
            'ƒ∞MES DEVIR', 'ƒ∞MES ALI≈û', 'ƒ∞MES SATIS', 'ƒ∞MES STOK',
            'ƒ∞Kƒ∞TELLƒ∞ DEVIR', 'ƒ∞Kƒ∞TELLƒ∞ ALI≈û', 'ƒ∞Kƒ∞TELLƒ∞ SATIS', 'ƒ∞Kƒ∞TELLƒ∞ STOK',
            'BOLU DEVIR', 'BOLU ALI≈û', 'BOLU SATIS', 'BOLU STOK',
            'ANKARA DEVIR', 'ANKARA ALI≈û', 'ANKARA SATIS', 'ANKARA STOK',
            # no2
            'not',
            # Depo Bakiye kolonlarƒ±
            'Maslak Depo Bakiye', 'Bolu Depo Bakiye', 'ƒ∞mes Depo Bakiye', 'Ankara Depo Bakiye', 'ƒ∞kitelli Depo Bakiye',
            # Kampanya Tipi
            'Kampanya Tipi',
            # Toplam ƒ∞sk
            'Toplam ƒ∞sk',
            # Toplam Depo Bakiye
            'Toplam Depo Bakiye',
            # Tedarik√ßi bakiye kolonlarƒ±
            'Maslak Tedarik√ßi Bakiye', 'Bolu Tedarik√ßi Bakiye', 'ƒ∞mes Tedarik√ßi Bakiye', 'Ankara Tedarik√ßi Bakiye', 'ƒ∞kitelli Tedarik√ßi Bakiye',
            # Paket Adetleri
            'Paket Adetleri',
            # Sipari≈ü kolonlarƒ±
            'Maslak Sipari≈ü', 'Bolu Sipari≈ü', 'ƒ∞mes Sipari≈ü', 'Ankara Sipari≈ü', 'ƒ∞kitelli Sipari≈ü',
            # Ay ba≈ülƒ±klarƒ±
            'Aƒüustos_1', 'Eyl√ºl_1', 'Aƒüustos_2', 'Eyl√ºl_2', 'Aƒüustos_3', 'Eyl√ºl_3', 'Aƒüustos_4', 'Eyl√ºl_4', 'Aƒüustos_5', 'Eyl√ºl_5',
            # Diƒüer s√ºtunlar
            'TOPL.FAT.ADT', 'M√ú≈ûT.SAY.', 'SATƒ±≈û FIYATƒ±', 'D√ñVIZ CINSI (S)', 'URUNKODU_3',
            # Son ba≈ülƒ±klar
            'Kampanya Tipi', 'not', 'ƒ∞SK', 'PRƒ∞M', 'B√úT√áE', 'liste', 'TD SF', 'Toplam ƒ∞sk', 'Net Fiyat Kampanyasƒ±'
        ]
        
        # Mevcut s√ºtunlarƒ± filtrele ve sƒ±rala
        available_cols = [col for col in desired_order if col in new_df.columns]
        if len(available_cols) > 0:
            new_df = new_df[available_cols]
        
        # Toplam Depo Bakiye hesaplama
        depo_bakiye_cols = ['Maslak Depo Bakiye', 'Bolu Depo Bakiye', 'ƒ∞mes Depo Bakiye', 'Ankara Depo Bakiye', 'ƒ∞kitelli Depo Bakiye']
        available_depo_cols = [col for col in depo_bakiye_cols if col in new_df.columns]
        
        if available_depo_cols and 'Toplam Depo Bakiye' in new_df.columns:
            # Sayƒ±sal deƒüerlere √ßevir ve topla
            for col in available_depo_cols:
                new_df[col] = pd.to_numeric(new_df[col], errors='coerce').fillna(0)
            
            # Toplam hesapla
            new_df['Toplam Depo Bakiye'] = new_df[available_depo_cols].sum(axis=1)
        
        return new_df
    
    except Exception as e:
        st.error(f"D√∂n√º≈ü√ºm hatasƒ±: {str(e)}")
        return pd.DataFrame()

@st.cache_data(show_spinner="Marka e≈üle≈ütirme yapƒ±lƒ±yor...", ttl=3600)
def match_brands_parallel(main_df, uploaded_files):
    """Paralel marka e≈üle≈ütirme"""
    try:
        # Marka-Excel e≈üle≈ütirme s√∂zl√ºƒü√º
        brand_excel_mapping = {
            'SCHAEFFLER LUK': 'excel1',
            'ZF ƒ∞THAL': 'excel2', 
            'DELPHI': 'excel3',
            'ZF YERLƒ∞': 'excel4',
            'VALEO': 'excel5',
            'FILTRON': 'excel6',
            'MANN': 'excel7'
        }
        
        # Ana DataFrame'i kopyala
        result_df = main_df.copy()
        
        # CAT4 kolonunu kontrol et
        if 'CAT4' not in main_df.columns:
            st.warning("CAT4 kolonu bulunamadƒ±!")
            return main_df
        
        # Paralel i≈üleme i√ßin marka verilerini topla
        brand_tasks = []
        for brand, excel_key in brand_excel_mapping.items():
            if excel_key in uploaded_files and uploaded_files[excel_key] is not None:
                brand_tasks.append((brand, uploaded_files[excel_key]))
        
        # Paralel marka verisi okuma
        brand_data = {}
        with ThreadPoolExecutor(max_workers=4) as executor:
            future_to_brand = {
                executor.submit(load_brand_data_parallel, file, brand): brand 
                for brand, file in brand_tasks
            }
            
            for future in as_completed(future_to_brand):
                brand_name, brand_df = future.result()
                brand_data[brand_name] = brand_df
                st.success(f"‚úÖ {brand_name} verisi y√ºklendi: {len(brand_df)} satƒ±r")
        
        # Her marka i√ßin i≈ülem yap
        for brand, brand_df in brand_data.items():
            if len(brand_df) > 0:
                # CAT4'te bu markayƒ± ara (esnek arama)
                search_terms = [brand]
                
                # Schaeffler i√ßin √∂zel arama terimleri
                if 'Schaeffler' in brand:
                    search_terms.extend(['Schaeffler', 'Schaeffler Luk', 'SchaefflerLuk', 'SCHAEFFLER LUK', 'SCHAEFFLER'])
                
                # ZF i√ßin √∂zel arama terimleri
                if 'ZF' in brand:
                    search_terms.extend(['LEMF√ñRDER', 'TRW', 'SACHS', 'LEMFORDER', 'TRW', 'SACHS'])
                
                # Mann i√ßin √∂zel arama terimleri
                if 'MANN' in brand:
                    search_terms.extend(['MANN', 'MANN FILTER', 'MANN-FILTER', 'MANNFILTER'])
                
                # Filtron i√ßin √∂zel arama terimleri
                if 'FILTRON' in brand:
                    search_terms.extend(['FILTRON', 'Fƒ∞LTRON', 'FILTRON FILTER', 'Fƒ∞LTRON Fƒ∞LTER'])
                
                # Debug: Arama terimlerini g√∂ster
                st.info(f"üîç {brand} i√ßin arama terimleri: {search_terms}")
                
                # T√ºm arama terimlerini dene
                brand_mask = pd.Series([False] * len(main_df))
                for search_term in search_terms:
                    temp_mask = main_df['CAT4'].str.contains(search_term, case=False, na=False)
                    brand_mask = brand_mask | temp_mask
                
                brand_count = brand_mask.sum()
                
                # Debug: CAT4'teki benzersiz deƒüerleri g√∂ster
                if brand_count == 0:
                    unique_cat4 = main_df['CAT4'].dropna().unique()
                    st.info(f"üîç CAT4 kolonundaki benzersiz deƒüerler: {list(unique_cat4[:10])}")
                
                if brand_count > 0:
                    st.info(f"üìä {brand} markasƒ± {brand_count} √ºr√ºn i√ßin bulundu")
                    
                    # Mann ve Filtron i√ßin normal i≈ülem (CAT4'te bulundu)
                    if ('MANN' in brand or 'FILTRON' in brand) and brand_count > 0:
                        st.info(f"üîÑ {brand} i√ßin normal i≈ülem yapƒ±lƒ±yor (CAT4'te bulundu)...")
                        # Burada normal i≈ülem yapƒ±lacak (Schaeffler gibi)
                        # ≈ûimdilik bo≈ü bƒ±rakƒ±yoruz, √∂zel i≈ülem kƒ±smƒ±nda yapƒ±lacak
                    
                    # Schaeffler Luk i√ßin tedarik√ßi bakiye i≈ülemi
                    if 'SCHAEFFLER LUK' in brand:
                        try:
                            # Tedarik√ßi bakiye kolonlarƒ±nƒ± olu≈ütur
                            if 'ƒ∞mes Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞mes Tedarik√ßi Bakiye'] = 0
                            if 'Ankara Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarik√ßi Bakiye'] = 0
                            if 'Bolu Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarik√ßi Bakiye'] = 0
                            if 'Maslak Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarik√ßi Bakiye'] = 0
                            if 'ƒ∞kitelli Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞kitelli Tedarik√ßi Bakiye'] = 0
                            
                            # Schaeffler verilerini i≈üle
                            schaeffler_df = brand_df.copy()
                            
                            # PO Number(L) kolonunu kontrol et
                            if 'PO Number(L)' in schaeffler_df.columns:
                                # Tedarik√ßi kodlarƒ±nƒ± belirle
                                schaeffler_df['Tedarik√ßi'] = schaeffler_df['PO Number(L)'].astype(str).apply(
                                    lambda x: 'ƒ∞mes' if 'IME' in x or '285' in x
                                    else 'Ankara' if 'ANK' in x or '321' in x
                                    else 'Bolu' if '322' in x
                                    else 'Maslak' if '323' in x
                                    else 'ƒ∞kitelli' if 'IKI' in x or '324' in x
                                    else 'Diƒüer'
                                )
                                
                                # Catalogue Number i≈üleme
                                if 'Catalogue number' in schaeffler_df.columns:
                                    # Sondaki 0'ƒ± sil ve bo≈üluklarƒ± temizle
                                    schaeffler_df['Catalogue_clean'] = schaeffler_df['Catalogue number'].astype(str).str.rstrip('0').str.strip()
                                
                                # Ordered Quantity kontrol√º
                                if 'Ordered quantity' in schaeffler_df.columns:
                                    # Tedarik√ßi bazƒ±nda grupla ve topla
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = schaeffler_df[schaeffler_df['Tedarik√ßi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Catalogue number bazƒ±nda topla
                                            grouped = tedarikci_data.groupby('Catalogue_clean')['Ordered quantity'].sum().reset_index()
                                            
                                            # Ana DataFrame ile e≈üle≈ütir
                                            for _, row in grouped.iterrows():
                                                catalogue_num = row['Catalogue_clean']
                                                quantity = row['Ordered quantity']
                                                
                                                # URUNKODU ile e≈üle≈ütir (bo≈üluklarƒ± temizle)
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip()
                                                match_mask = urunkodu_clean.str.contains(catalogue_num, case=False, na=False)
                                                
                                                if match_mask.sum() > 0:
                                                    # Tedarik√ßi kolonunu g√ºncelle (toplama ile)
                                                    if tedarikci == 'ƒ∞mes':
                                                        result_df.loc[match_mask, 'ƒ∞mes Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'ƒ∞kitelli':
                                                        result_df.loc[match_mask, 'ƒ∞kitelli Tedarik√ßi Bakiye'] += quantity
                                
                                st.success(f"‚úÖ Schaeffler Luk verileri i≈ülendi: {len(schaeffler_df)} satƒ±r")
                            else:
                                st.warning("‚ö†Ô∏è Schaeffler dosyasƒ±nda 'PO Number(L)' kolonu bulunamadƒ±")
                                
                        except Exception as e:
                            st.error(f"‚ùå Schaeffler veri i≈üleme hatasƒ±: {str(e)}")
                    
                    # ZF ƒ∞thal i√ßin tedarik√ßi bakiye i≈ülemi
                    elif 'ZF ƒ∞THAL' in brand:
                        try:
                            # Tedarik√ßi bakiye kolonlarƒ±nƒ± olu≈ütur
                            if 'ƒ∞mes Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞mes Tedarik√ßi Bakiye'] = 0
                            if 'Ankara Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarik√ßi Bakiye'] = 0
                            if 'Bolu Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarik√ßi Bakiye'] = 0
                            if 'Maslak Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarik√ßi Bakiye'] = 0
                            if 'ƒ∞kitelli Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞kitelli Tedarik√ßi Bakiye'] = 0
                            
                            # ZF ƒ∞thal verilerini i≈üle
                            zf_ithal_df = brand_df.copy()
                            
                            # Material kolonunu kontrol et
                            if 'Material' in zf_ithal_df.columns:
                                # Material kodunu i≈üle - d√ºzeltilmi≈ü kural
                                zf_ithal_df['Material_clean'] = zf_ithal_df['Material'].astype(str).apply(
                                    lambda x: x.split(':')[1].replace(' ', '') if ':' in x and (x.startswith('LF:') or x.startswith('SX:'))  # LF: veya SX: ile ba≈ülƒ±yorsa : sonrasƒ±nƒ± al
                                    else x.split(':')[0].strip() if ':' in x and not (x.startswith('LF:') or x.startswith('SX:'))  # Diƒüerlerinde : √∂ncesini al
                                    else x.replace(' ', '')  # : yoksa bo≈üluklarƒ± sil
                                )
                                
                                # Debug: ƒ∞lk 5 √∂rnek g√∂ster
                                st.info(f"üîç ZF ƒ∞thal Material i≈üleme √∂rnekleri:")
                                for i, (original, cleaned) in enumerate(zip(zf_ithal_df['Material'].head(), zf_ithal_df['Material_clean'].head())):
                                    st.write(f"  {original} ‚Üí {cleaned}")
                                
                                # Debug: E≈üle≈ütirme √∂rnekleri
                                st.info(f"üîç ZF ƒ∞thal e≈üle≈ütirme √∂rnekleri:")
                                for i, material_num in enumerate(zf_ithal_df['Material_clean'].head()):
                                    urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip()
                                    duzenlenmis_clean = result_df['D√ºzenlenmi≈ü √úr√ºn Kodu'].astype(str).str.replace(' ', '', regex=False)
                                    
                                    match_urun = urunkodu_clean.str.contains(material_num, case=False, na=False).sum()
                                    match_duzen = duzenlenmis_clean.str.contains(material_num, case=False, na=False).sum()
                                    
                                    st.write(f"  {material_num} ‚Üí URUNKODU: {match_urun}, D√ºzenlenmi≈ü: {match_duzen}")
                                
                                # Purchase order no. kolonunu kontrol et
                                if 'Purchase order no.' in zf_ithal_df.columns:
                                    # Tedarik√ßi kodlarƒ±nƒ± belirle
                                    zf_ithal_df['Tedarik√ßi'] = zf_ithal_df['Purchase order no.'].astype(str).apply(
                                        lambda x: 'ƒ∞mes' if 'IME' in x or '285' in x or 'ƒ∞ST' in x or 'IST' in x
                                        else 'Ankara' if 'ANK' in x or '321' in x
                                        else 'Bolu' if '322' in x
                                        else 'Maslak' if '323' in x
                                        else 'ƒ∞kitelli' if 'IKI' in x or '324' in x
                                        else 'Diƒüer'
                                    )
                                    
                                    # Debug: Tedarik√ßi daƒüƒ±lƒ±mƒ±nƒ± g√∂ster
                                    tedarikci_counts = zf_ithal_df['Tedarik√ßi'].value_counts()
                                    st.info(f"üîç ZF ƒ∞thal Tedarik√ßi daƒüƒ±lƒ±mƒ±:")
                                    for tedarikci, count in tedarikci_counts.items():
                                        st.write(f"  {tedarikci}: {count} satƒ±r")
                                    
                                    # Debug: √ñrnek Purchase order no. kodlarƒ± g√∂ster
                                    st.info("üîç ZF ƒ∞thal Purchase order no. √∂rnekleri:")
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = zf_ithal_df[zf_ithal_df['Tedarik√ßi'] == tedarikci]
                                        if len(tedarikci_data) > 0:
                                            sample_codes = tedarikci_data['Purchase order no.'].head(3).tolist()
                                            st.write(f"  {tedarikci}: {sample_codes}")
                                
                                # Qty.in Del. ve Open quantity kolonlarƒ±nƒ± kontrol et
                                if 'Qty.in Del.' in zf_ithal_df.columns and 'Open quantity' in zf_ithal_df.columns:
                                    # Tedarik√ßi bazƒ±nda grupla ve topla
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = zf_ithal_df[zf_ithal_df['Tedarik√ßi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Material_clean bazƒ±nda topla
                                            grouped = tedarikci_data.groupby('Material_clean').agg({
                                                'Qty.in Del.': 'sum',
                                                'Open quantity': 'sum'
                                            }).reset_index()
                                            
                                            # Ana DataFrame ile e≈üle≈ütir (LPR, Lemforder, TRW markalarƒ±)
                                            for _, row in grouped.iterrows():
                                                material_num = row['Material_clean']
                                                qty_del = row['Qty.in Del.']
                                                open_qty = row['Open quantity']
                                                total_qty = qty_del + open_qty
                                                
                                                # LEMF√ñRDER, TRW, SACHS markalarƒ±nƒ± ara
                                                lemforder_mask = result_df['CAT4'].str.contains('LEMF√ñRDER', case=False, na=False)
                                                trw_mask = result_df['CAT4'].str.contains('TRW', case=False, na=False)
                                                sachs_mask = result_df['CAT4'].str.contains('SACHS', case=False, na=False)
                                                
                                                # Hem URUNKODU hem de D√ºzenlenmi≈ü √úr√ºn Kodu ile e≈üle≈ütir
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip()
                                                duzenlenmis_clean = result_df['D√ºzenlenmi≈ü √úr√ºn Kodu'].astype(str).str.replace(' ', '', regex=False)
                                                
                                                # Bo≈üluklarƒ± temizlenmi≈ü versiyonlar da olu≈ütur
                                                urunkodu_no_space = urunkodu_clean.str.replace(' ', '', regex=False)
                                                duzenlenmis_no_space = duzenlenmis_clean.str.replace(' ', '', regex=False)
                                                material_no_space = material_num.replace(' ', '')
                                                
                                                # D√∂rt farklƒ± e≈üle≈ütirme y√∂ntemi dene
                                                match_mask_urun = urunkodu_clean.str.contains(material_num, case=False, na=False)
                                                match_mask_duzen = duzenlenmis_clean.str.contains(material_num, case=False, na=False)
                                                match_mask_urun_no_space = urunkodu_no_space.str.contains(material_no_space, case=False, na=False)
                                                match_mask_duzen_no_space = duzenlenmis_no_space.str.contains(material_no_space, case=False, na=False)
                                                match_mask = match_mask_urun | match_mask_duzen | match_mask_urun_no_space | match_mask_duzen_no_space
                                                
                                                # LEMF√ñRDER, TRW, SACHS markalarƒ± ile birle≈ütir
                                                final_mask = match_mask & (lemforder_mask | trw_mask | sachs_mask)
                                                
                                                if final_mask.sum() > 0:
                                                    # Tedarik√ßi kolonunu g√ºncelle (toplama ile)
                                                    if tedarikci == 'ƒ∞mes':
                                                        result_df.loc[final_mask, 'ƒ∞mes Tedarik√ßi Bakiye'] += total_qty
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[final_mask, 'Ankara Tedarik√ßi Bakiye'] += total_qty
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[final_mask, 'Bolu Tedarik√ßi Bakiye'] += total_qty
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[final_mask, 'Maslak Tedarik√ßi Bakiye'] += total_qty
                                                    elif tedarikci == 'ƒ∞kitelli':
                                                        result_df.loc[final_mask, 'ƒ∞kitelli Tedarik√ßi Bakiye'] += total_qty
                                                    
                                                    # Debug bilgisi
                                                    st.info(f"üîç ZF ƒ∞thal: {material_num} ‚Üí {final_mask.sum()} e≈üle≈üme bulundu")
                                                else:
                                                    # E≈üle≈üme bulunamadƒ±ƒüƒ±nda detaylƒ± debug bilgisi
                                                    st.warning(f"‚ö†Ô∏è ZF ƒ∞thal: {material_num} i√ßin e≈üle≈üme bulunamadƒ±")
                                                    st.write(f"  Material (temiz): {material_num}")
                                                    st.write(f"  Material (bo≈üluksuz): {material_no_space}")
                                                    
                                                    # √ñrnek URUNKODU ve D√ºzenlenmi≈ü √úr√ºn Kodu g√∂ster
                                                    sample_urun = result_df['URUNKODU'].head(5).tolist()
                                                    sample_duzen = result_df['D√ºzenlenmi≈ü √úr√ºn Kodu'].head(5).tolist()
                                                    st.write(f"  √ñrnek URUNKODU: {sample_urun}")
                                                    st.write(f"  √ñrnek D√ºzenlenmi≈ü: {sample_duzen}")
                                
                                st.success(f"‚úÖ ZF ƒ∞thal verileri i≈ülendi: {len(zf_ithal_df)} satƒ±r")
                            else:
                                st.warning("‚ö†Ô∏è ZF ƒ∞thal dosyasƒ±nda 'Material' kolonu bulunamadƒ±")
                                
                        except Exception as e:
                            st.error(f"‚ùå ZF ƒ∞thal veri i≈üleme hatasƒ±: {str(e)}")
                
                    # ZF Yerli i√ßin tedarik√ßi bakiye i≈ülemi
                    elif 'ZF YERLƒ∞' in brand:
                        try:
                            # Tedarik√ßi bakiye kolonlarƒ±nƒ± olu≈ütur
                            if 'ƒ∞mes Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞mes Tedarik√ßi Bakiye'] = 0
                            if 'Ankara Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarik√ßi Bakiye'] = 0
                            if 'Bolu Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarik√ßi Bakiye'] = 0
                            if 'Maslak Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarik√ßi Bakiye'] = 0
                            if 'ƒ∞kitelli Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞kitelli Tedarik√ßi Bakiye'] = 0
                            
                            # ZF Yerli verilerini i≈üle
                            zf_yerli_df = brand_df.copy()
                            
                            # Basic No. kolonunu kontrol et
                            if 'Basic No.' in zf_yerli_df.columns:
                                # Basic No. kodunu temizle
                                zf_yerli_df['Basic_clean'] = zf_yerli_df['Basic No.'].astype(str).str.strip()
                                
                                # Ship-to Name kolonunu kontrol et
                                if 'Ship-to Name' in zf_yerli_df.columns:
                                    # Tedarik√ßi kodlarƒ±nƒ± belirle
                                    zf_yerli_df['Tedarik√ßi'] = zf_yerli_df['Ship-to Name'].astype(str).apply(
                                        lambda x: 'ƒ∞mes' if 'IME' in x or '285' in x or 'IST' in x or 'ƒ∞ST' in x
                                        else 'Ankara' if 'ANK' in x or '321' in x
                                        else 'Bolu' if '322' in x
                                        else 'Maslak' if '323' in x
                                        else 'ƒ∞kitelli' if 'IKI' in x or '324' in x
                                        else 'Diƒüer'
                                    )
                                    
                                    # Debug: Tedarik√ßi daƒüƒ±lƒ±mƒ±nƒ± g√∂ster
                                    tedarikci_counts = zf_yerli_df['Tedarik√ßi'].value_counts()
                                    st.info(f"üîç ZF Yerli Tedarik√ßi daƒüƒ±lƒ±mƒ±:")
                                    for tedarikci, count in tedarikci_counts.items():
                                        st.write(f"  {tedarikci}: {count} satƒ±r")
                                
                                # Outstanding Quantity kolonunu kontrol et
                                if 'Outstanding Quantity' in zf_yerli_df.columns:
                                    # Tedarik√ßi bazƒ±nda grupla ve topla
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = zf_yerli_df[zf_yerli_df['Tedarik√ßi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Basic_clean bazƒ±nda topla
                                            grouped = tedarikci_data.groupby('Basic_clean')['Outstanding Quantity'].sum().reset_index()
                                            
                                            # Ana DataFrame ile e≈üle≈ütir (D√ºzenlenmi≈ü √úr√ºn Kodu ile)
                                            for _, row in grouped.iterrows():
                                                basic_num = row['Basic_clean']
                                                quantity = row['Outstanding Quantity']
                                                
                                                # LEMF√ñRDER, TRW, SACHS markalarƒ±nƒ± ara
                                                lemforder_mask = result_df['CAT4'].str.contains('LEMF√ñRDER', case=False, na=False)
                                                trw_mask = result_df['CAT4'].str.contains('TRW', case=False, na=False)
                                                sachs_mask = result_df['CAT4'].str.contains('SACHS', case=False, na=False)
                                                
                                                # D√ºzenlenmi≈ü √úr√ºn Kodu ile e≈üle≈ütir (bo≈üluklarƒ± temizle)
                                                duzenlenmis_clean = result_df['D√ºzenlenmi≈ü √úr√ºn Kodu'].astype(str).str.strip()
                                                match_mask = duzenlenmis_clean.str.contains(basic_num, case=False, na=False)
                                                
                                                # LEMF√ñRDER, TRW, SACHS markalarƒ± ile birle≈ütir
                                                final_mask = match_mask & (lemforder_mask | trw_mask | sachs_mask)
                                                
                                                if final_mask.sum() > 0:
                                                    # Tedarik√ßi kolonunu g√ºncelle (toplama ile)
                                                    if tedarikci == 'ƒ∞mes':
                                                        result_df.loc[final_mask, 'ƒ∞mes Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[final_mask, 'Ankara Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[final_mask, 'Bolu Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[final_mask, 'Maslak Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'ƒ∞kitelli':
                                                        result_df.loc[final_mask, 'ƒ∞kitelli Tedarik√ßi Bakiye'] += quantity
                                
                                st.success(f"‚úÖ ZF Yerli verileri i≈ülendi: {len(zf_yerli_df)} satƒ±r")
                            else:
                                st.warning("‚ö†Ô∏è ZF Yerli dosyasƒ±nda 'Basic No.' kolonu bulunamadƒ±")
                                
                        except Exception as e:
                            st.error(f"‚ùå ZF Yerli veri i≈üleme hatasƒ±: {str(e)}")
                
                    # Valeo i√ßin tedarik√ßi bakiye i≈ülemi
                    elif 'VALEO' in brand:
                        try:
                            # Tedarik√ßi bakiye kolonlarƒ±nƒ± olu≈ütur
                            if 'ƒ∞mes Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞mes Tedarik√ßi Bakiye'] = 0
                            if 'Ankara Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarik√ßi Bakiye'] = 0
                            if 'Bolu Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarik√ßi Bakiye'] = 0
                            if 'Maslak Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarik√ßi Bakiye'] = 0
                            if 'ƒ∞kitelli Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞kitelli Tedarik√ßi Bakiye'] = 0
                            
                            # Valeo verilerini i≈üle
                            valeo_df = brand_df.copy()
                            
                            # M√º≈üteri P/O No. kolonunu kontrol et
                            if 'M√º≈üteri P/O No.' in valeo_df.columns:
                                # Tedarik√ßi kodlarƒ±nƒ± belirle
                                valeo_df['Tedarik√ßi'] = valeo_df['M√º≈üteri P/O No.'].astype(str).apply(
                                    lambda x: 'ƒ∞mes' if 'IME' in x or '285' in x
                                    else 'Ankara' if 'ANK' in x or '321' in x
                                    else 'Bolu' if '322' in x
                                    else 'Maslak' if '323' in x
                                    else 'ƒ∞kitelli' if 'IKI' in x or '324' in x
                                    else 'Diƒüer'
                                )
                                
                                # Valeo Ref. kolonunu kontrol et
                                if 'Valeo Ref.' in valeo_df.columns:
                                    # Valeo Ref. kodunu temizle
                                    valeo_df['Valeo_clean'] = valeo_df['Valeo Ref.'].astype(str).str.strip()
                                
                                # Sipari≈ü Adeti kolonunu kontrol et
                                if 'Sipari≈ü Adeti' in valeo_df.columns:
                                    # Tedarik√ßi bazƒ±nda grupla ve topla
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = valeo_df[valeo_df['Tedarik√ßi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Valeo_clean bazƒ±nda topla
                                            grouped = tedarikci_data.groupby('Valeo_clean')['Sipari≈ü Adeti'].sum().reset_index()
                                            
                                            # Ana DataFrame ile e≈üle≈ütir
                                            for _, row in grouped.iterrows():
                                                valeo_ref = row['Valeo_clean']
                                                quantity = row['Sipari≈ü Adeti']
                                                
                                                # URUNKODU ile e≈üle≈ütir (bo≈üluklarƒ± temizle)
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip()
                                                match_mask = urunkodu_clean.str.contains(valeo_ref, case=False, na=False)
                                                
                                                if match_mask.sum() > 0:
                                                    # Tedarik√ßi kolonunu g√ºncelle (toplama ile)
                                                    if tedarikci == 'ƒ∞mes':
                                                        result_df.loc[match_mask, 'ƒ∞mes Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'ƒ∞kitelli':
                                                        result_df.loc[match_mask, 'ƒ∞kitelli Tedarik√ßi Bakiye'] += quantity
                                
                                st.success(f"‚úÖ Valeo verileri i≈ülendi: {len(valeo_df)} satƒ±r")
                            else:
                                st.warning("‚ö†Ô∏è Valeo dosyasƒ±nda 'M√º≈üteri P/O No.' kolonu bulunamadƒ±")
                                
                        except Exception as e:
                            st.error(f"‚ùå Valeo veri i≈üleme hatasƒ±: {str(e)}")
                
                    # Delphi i√ßin tedarik√ßi bakiye i≈ülemi
                    elif 'DELPHI' in brand:
                        try:
                            # Tedarik√ßi bakiye kolonlarƒ±nƒ± olu≈ütur
                            if 'ƒ∞mes Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞mes Tedarik√ßi Bakiye'] = 0
                            if 'Ankara Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarik√ßi Bakiye'] = 0
                            if 'Bolu Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarik√ßi Bakiye'] = 0
                            if 'Maslak Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarik√ßi Bakiye'] = 0
                            if 'ƒ∞kitelli Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞kitelli Tedarik√ßi Bakiye'] = 0
                            
                            # Delphi verilerini i≈üle
                            delphi_df = brand_df.copy()
                            
                            # ≈ûube kolonunu kontrol et
                            if '≈ûube' in delphi_df.columns:
                                # Tedarik√ßi kodlarƒ±nƒ± belirle
                                delphi_df['Tedarik√ßi'] = delphi_df['≈ûube'].astype(str).apply(
                                    lambda x: 'Bolu' if 'Teknik Dizel-Bolu' in x
                                    else 'ƒ∞mes' if 'Teknik Dizel-√úmraniye' in x
                                    else 'Maslak' if 'Teknik Dizel-Maslak' in x
                                    else 'Ankara' if 'Teknik Dizel-Ankara' in x
                                    else 'ƒ∞kitelli' if 'Teknik Dizel-ƒ∞kitelli' in x
                                    else 'Diƒüer'
                                )
                                
                                # Material kolonunu kontrol et
                                if 'Material' in delphi_df.columns:
                                    # Material kodunu temizle
                                    delphi_df['Material_clean'] = delphi_df['Material'].astype(str).str.strip()
                                
                                # Cum.qty kolonunu kontrol et
                                if 'Cum.qty' in delphi_df.columns:
                                    # Tedarik√ßi bazƒ±nda grupla ve topla
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = delphi_df[delphi_df['Tedarik√ßi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Material_clean bazƒ±nda topla
                                            grouped = tedarikci_data.groupby('Material_clean')['Cum.qty'].sum().reset_index()
                                            
                                            # Ana DataFrame ile e≈üle≈ütir
                                            for _, row in grouped.iterrows():
                                                material_num = row['Material_clean']
                                                quantity = row['Cum.qty']
                                                
                                                # URUNKODU ile e≈üle≈ütir (bo≈üluklarƒ± temizle)
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip()
                                                match_mask = urunkodu_clean.str.contains(material_num, case=False, na=False)
                                                
                                                if match_mask.sum() > 0:
                                                    # Tedarik√ßi kolonunu g√ºncelle (toplama ile)
                                                    if tedarikci == 'ƒ∞mes':
                                                        result_df.loc[match_mask, 'ƒ∞mes Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarik√ßi Bakiye'] += quantity
                                                    elif tedarikci == 'ƒ∞kitelli':
                                                        result_df.loc[match_mask, 'ƒ∞kitelli Tedarik√ßi Bakiye'] += quantity
                                
                                st.success(f"‚úÖ Delphi verileri i≈ülendi: {len(delphi_df)} satƒ±r")
                            else:
                                st.warning("‚ö†Ô∏è Delphi dosyasƒ±nda '≈ûube' kolonu bulunamadƒ±")
                                
                        except Exception as e:
                            st.error(f"‚ùå Delphi veri i≈üleme hatasƒ±: {str(e)}")
                
                    # Mann ve Filtron i√ßin tedarik√ßi bakiye i≈ülemi
                    if 'MANN' in brand or 'Fƒ∞LTRON' in brand:
                        try:
                            # Tedarik√ßi bakiye kolonlarƒ±nƒ± olu≈ütur
                            if 'ƒ∞mes Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞mes Tedarik√ßi Bakiye'] = 0
                            if 'Ankara Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarik√ßi Bakiye'] = 0
                            if 'Bolu Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarik√ßi Bakiye'] = 0
                            if 'Maslak Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarik√ßi Bakiye'] = 0
                            if 'ƒ∞kitelli Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞kitelli Tedarik√ßi Bakiye'] = 0
                            
                            # Mann/Filtron verilerini i≈üle
                            brand_df_processed = brand_df.copy()
                            
                            # Material Adƒ± kolonunu kontrol et (farklƒ± isimler i√ßin)
                            material_col = None
                            for col_name in ['Material Adƒ±', 'Material', 'Material Name', '√úr√ºn Kodu', 'Product Code', 'Material Kodu', 'Malzeme Kodu', 'Malzeme Adƒ±']:
                                if col_name in brand_df_processed.columns:
                                    material_col = col_name
                                    break
                            
                            if material_col:
                                st.info(f"üîç {brand} i√ßin {material_col} kolonu bulundu!")
                                
                                # Material kodunu temizle (bulunan kolon adƒ±nƒ± kullan)
                                brand_df_processed['Material_clean'] = brand_df_processed[material_col].astype(str).str.strip()
                                
                                # M√º≈üteri Satƒ±nAlma No kolonunu kontrol et
                                if 'M√º≈üteri Satƒ±nAlma No' in brand_df_processed.columns:
                                    # Tedarik√ßi kodlarƒ±nƒ± belirle
                                    brand_df_processed['Tedarik√ßi'] = brand_df_processed['M√º≈üteri Satƒ±nAlma No'].astype(str).apply(
                                        lambda x: 'Ankara' if 'AAS' in x
                                        else 'ƒ∞mes' if 'DAS' in x
                                        else 'Bolu' if 'BAS' in x
                                        else 'Maslak' if 'MAS' in x
                                        else 'ƒ∞kitelli' if 'EAS' in x
                                        else 'Diƒüer'
                                    )
                                
                                # A√ßƒ±k Sipari≈ü Adedi kolonunu kontrol et
                                if 'A√ßƒ±k Sipari≈ü Adedi' in brand_df_processed.columns:
                                    # Tedarik√ßi bazƒ±nda grupla ve topla
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = brand_df_processed[brand_df_processed['Tedarik√ßi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Material_clean bazƒ±nda topla
                                            grouped = tedarikci_data.groupby('Material_clean')['A√ßƒ±k Sipari≈ü Adedi'].sum().reset_index()
                                            
                                            # Ana DataFrame ile e≈üle≈ütir
                                            for _, row in grouped.iterrows():
                                                material_num = row['Material_clean']
                                                quantity = row['A√ßƒ±k Sipari≈ü Adedi']
                                                
                                                # Hem URUNKODU hem de D√ºzenlenmi≈ü √úr√ºn Kodu ile e≈üle≈ütir
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip()
                                                duzenlenmis_clean = result_df['D√ºzenlenmi≈ü √úr√ºn Kodu'].astype(str).str.replace(' ', '', regex=False)
                                                material_clean_no_space = material_num.replace(' ', '')
                                                
                                                # ƒ∞ki farklƒ± e≈üle≈ütirme y√∂ntemi dene
                                                match_mask_urun = urunkodu_clean.str.contains(material_clean_no_space, case=False, na=False)
                                                match_mask_duzen = duzenlenmis_clean.str.contains(material_clean_no_space, case=False, na=False)
                                                match_mask = match_mask_urun | match_mask_duzen
                                                
                                                if match_mask.sum() > 0:
                                                    # Tedarik√ßi kolonunu g√ºncelle (toplama ile)
                                                    if tedarikci == 'ƒ∞mes':
                                                        result_df.loc[match_mask, 'ƒ∞mes Tedarik√ßi Bakiye'] += quantity
                                                        st.success(f"‚úÖ {brand} {tedarikci}: {material_num} ‚Üí {quantity} adet eklendi")
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarik√ßi Bakiye'] += quantity
                                                        st.success(f"‚úÖ {brand} {tedarikci}: {material_num} ‚Üí {quantity} adet eklendi")
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarik√ßi Bakiye'] += quantity
                                                        st.success(f"‚úÖ {brand} {tedarikci}: {material_num} ‚Üí {quantity} adet eklendi")
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarik√ßi Bakiye'] += quantity
                                                        st.success(f"‚úÖ {brand} {tedarikci}: {material_num} ‚Üí {quantity} adet eklendi")
                                                    elif tedarikci == 'ƒ∞kitelli':
                                                        result_df.loc[match_mask, 'ƒ∞kitelli Tedarik√ßi Bakiye'] += quantity
                                                        st.success(f"‚úÖ {brand} {tedarikci}: {material_num} ‚Üí {quantity} adet eklendi")
                                                else:
                                                    # E≈üle≈üme bulunamadƒ±ƒüƒ±nda debug bilgisi
                                                    st.warning(f"‚ö†Ô∏è {brand}: {material_num} i√ßin e≈üle≈üme bulunamadƒ±")
                                
                                st.success(f"‚úÖ {brand} verileri i≈ülendi: {len(brand_df_processed)} satƒ±r")
                                
                                # Debug: Sonu√ß kontrol√º
                                st.info(f"üîç {brand} sonu√ß kontrol√º:")
                                for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                    col_name = f"{tedarikci} Tedarik√ßi Bakiye"
                                    if col_name in result_df.columns:
                                        total = result_df[col_name].sum()
                                        st.write(f"  {tedarikci}: {total} adet")
                            else:
                                st.warning(f"‚ö†Ô∏è {brand} dosyasƒ±nda gerekli kolonlar bulunamadƒ±. Mevcut kolonlar: {list(brand_df_processed.columns)}")
                                
                        except Exception as e:
                            st.error(f"‚ùå {brand} veri i≈üleme hatasƒ±: {str(e)}")
                
                else:
                    st.warning(f"‚ö†Ô∏è {brand} markasƒ± CAT4 kolonunda bulunamadƒ±")
                            # Tedarik√ßi bakiye kolonlarƒ±nƒ± olu≈ütur
                            if 'ƒ∞mes Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞mes Tedarik√ßi Bakiye'] = 0
                            if 'Ankara Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Ankara Tedarik√ßi Bakiye'] = 0
                            if 'Bolu Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Bolu Tedarik√ßi Bakiye'] = 0
                            if 'Maslak Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['Maslak Tedarik√ßi Bakiye'] = 0
                            if 'ƒ∞kitelli Tedarik√ßi Bakiye' not in result_df.columns:
                                result_df['ƒ∞kitelli Tedarik√ßi Bakiye'] = 0
                            
                            # Mann/Filtron verilerini i≈üle
                            brand_df_processed = brand_df.copy()
                            
                            # Material Adƒ± kolonunu kontrol et (farklƒ± isimler i√ßin)
                            material_col = None
                            for col_name in ['Material Adƒ±', 'Material', 'Material Name', '√úr√ºn Kodu', 'Product Code', 'Material Kodu', 'Malzeme Kodu', 'Malzeme Adƒ±']:
                                if col_name in brand_df_processed.columns:
                                    material_col = col_name
                                    break
                            

                            
                            # Debug: T√ºm kolonlarƒ± g√∂ster
                            st.info(f"üîç {brand} dosyasƒ±ndaki t√ºm kolonlar:")
                            for i, col in enumerate(brand_df_processed.columns):
                                st.write(f"  {i+1}. {col}")
                            
                            # √ñnemli kolonlarƒ± kontrol et
                            important_cols = ['M√º≈üteri Satƒ±nAlma No', 'A√ßƒ±k Sipari≈ü Adedi', 'Material Kodu', 'Material Adƒ±']
                            st.info(f"üîç {brand} √∂nemli kolonlar kontrol√º:")
                            for col in important_cols:
                                if col in brand_df_processed.columns:
                                    st.success(f"  ‚úÖ {col} - BULUNDU")
                                else:
                                    st.error(f"  ‚ùå {col} - BULUNAMADI")
                            
                            if material_col:
                                # Material kodunu temizle
                                brand_df_processed['Material_clean'] = brand_df_processed[material_col].astype(str).str.strip()
                                
                                # Debug: Material kolonu √∂rnekleri g√∂ster
                                st.info(f"üîç {brand} Material kolonu √∂rnekleri:")
                                sample_materials = brand_df_processed[material_col].head(10).tolist()
                                for i, material in enumerate(sample_materials):
                                    st.write(f"  {i+1}. {material}")
                                
                                # Debug: Material_clean √∂rnekleri g√∂ster
                                st.info(f"üîç {brand} Material_clean √∂rnekleri:")
                                sample_cleaned = brand_df_processed['Material_clean'].head(10).tolist()
                                for i, cleaned in enumerate(sample_cleaned):
                                    st.write(f"  {i+1}. {cleaned}")
                                
                                # M√º≈üteri Satƒ±nAlma No kolonunu kontrol et
                                if 'M√º≈üteri Satƒ±nAlma No' in brand_df_processed.columns:
                                    # Tedarik√ßi kodlarƒ±nƒ± belirle
                                    brand_df_processed['Tedarik√ßi'] = brand_df_processed['M√º≈üteri Satƒ±nAlma No'].astype(str).apply(
                                        lambda x: 'Ankara' if 'AAS' in x
                                        else 'ƒ∞mes' if 'DAS' in x
                                        else 'Bolu' if 'BAS' in x
                                        else 'Maslak' if 'MAS' in x
                                        else 'ƒ∞kitelli' if 'EAS' in x
                                        else 'Diƒüer'
                                    )
                                
                                # A√ßƒ±k Sipari≈ü Adedi kolonunu kontrol et
                                if 'A√ßƒ±k Sipari≈ü Adedi' in brand_df_processed.columns:
                                    # Tedarik√ßi bazƒ±nda grupla ve topla
                                    for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                        tedarikci_data = brand_df_processed[brand_df_processed['Tedarik√ßi'] == tedarikci]
                                        
                                        if len(tedarikci_data) > 0:
                                            # Material_clean bazƒ±nda topla
                                            grouped = tedarikci_data.groupby('Material_clean')['A√ßƒ±k Sipari≈ü Adedi'].sum().reset_index()
                                            
                                            # Ana DataFrame ile e≈üle≈ütir
                                            for _, row in grouped.iterrows():
                                                material_num = row['Material_clean']
                                                quantity = row['A√ßƒ±k Sipari≈ü Adedi']
                                                
                                                # Hem URUNKODU hem de D√ºzenlenmi≈ü √úr√ºn Kodu ile e≈üle≈ütir
                                                urunkodu_clean = result_df['URUNKODU'].astype(str).str.strip()
                                                duzenlenmis_clean = result_df['D√ºzenlenmi≈ü √úr√ºn Kodu'].astype(str).str.replace(' ', '', regex=False)
                                                material_clean_no_space = material_num.replace(' ', '')
                                                
                                                # ƒ∞ki farklƒ± e≈üle≈ütirme y√∂ntemi dene
                                                match_mask_urun = urunkodu_clean.str.contains(material_clean_no_space, case=False, na=False)
                                                match_mask_duzen = duzenlenmis_clean.str.contains(material_clean_no_space, case=False, na=False)
                                                match_mask = match_mask_urun | match_mask_duzen
                                                
                                                if match_mask.sum() > 0:
                                                    # Tedarik√ßi kolonunu g√ºncelle
                                                    if tedarikci == 'ƒ∞mes':
                                                        result_df.loc[match_mask, 'ƒ∞mes Tedarik√ßi Bakiye'] = quantity
                                                    elif tedarikci == 'Ankara':
                                                        result_df.loc[match_mask, 'Ankara Tedarik√ßi Bakiye'] = quantity
                                                    elif tedarikci == 'Bolu':
                                                        result_df.loc[match_mask, 'Bolu Tedarik√ßi Bakiye'] = quantity
                                                    elif tedarikci == 'Maslak':
                                                        result_df.loc[match_mask, 'Maslak Tedarik√ßi Bakiye'] = quantity
                                                    elif tedarikci == 'ƒ∞kitelli':
                                                        result_df.loc[match_mask, 'ƒ∞kitelli Tedarik√ßi Bakiye'] = quantity
                                                    
                                                    # Debug bilgisi
                                                    st.info(f"üîç {brand}: {material_num} ‚Üí {match_mask.sum()} e≈üle≈üme bulundu")
                                                else:
                                                    # E≈üle≈üme bulunamadƒ±ƒüƒ±nda debug bilgisi
                                                    st.warning(f"‚ö†Ô∏è {brand}: {material_num} i√ßin e≈üle≈üme bulunamadƒ±")
                                
                                st.success(f"‚úÖ {brand} verileri i≈ülendi: {len(brand_df_processed)} satƒ±r")
                                
                                # Debug: Sonu√ß kontrol√º
                                st.info(f"üîç {brand} sonu√ß kontrol√º:")
                                for tedarikci in ['ƒ∞mes', 'Ankara', 'Bolu', 'Maslak', 'ƒ∞kitelli']:
                                    col_name = f"{tedarikci} Tedarik√ßi Bakiye"
                                    if col_name in result_df.columns:
                                        total = result_df[col_name].sum()
                                        st.write(f"  {tedarikci}: {total} adet")
                            else:
                                st.warning(f"‚ö†Ô∏è {brand} dosyasƒ±nda gerekli kolonlar bulunamadƒ±. Mevcut kolonlar: {list(brand_df_processed.columns)}")
                                
                        except Exception as e:
                            st.error(f"‚ùå {brand} veri i≈üleme hatasƒ±: {str(e)}")
        
        # Marka e≈üle≈ütirme sonrasƒ± toplam depo bakiyesi g√ºncelleme
        depo_bakiye_cols = ['Maslak Depo Bakiye', 'Bolu Depo Bakiye', 'ƒ∞mes Depo Bakiye', 'Ankara Depo Bakiye', 'ƒ∞kitelli Depo Bakiye']
        available_depo_cols = [col for col in depo_bakiye_cols if col in result_df.columns]
        
        if available_depo_cols and 'Toplam Depo Bakiye' in result_df.columns:
            # Sayƒ±sal deƒüerlere √ßevir ve topla
            for col in available_depo_cols:
                result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0)
            
            # Toplam hesapla
            result_df['Toplam Depo Bakiye'] = result_df[available_depo_cols].sum(axis=1)
            
            st.success(f"‚úÖ Toplam Depo Bakiye hesaplandƒ±: {len(available_depo_cols)} depo kolonu toplandƒ±")
        
        return result_df
        
    except Exception as e:
        st.error(f"Marka e≈üle≈ütirme hatasƒ±: {str(e)}")
        return main_df

@st.cache_data(show_spinner="Excel olu≈üturuluyor...", ttl=1800)
def format_excel_ultra_fast(df):
    """Ultra hƒ±zlƒ± Excel olu≈üturma - performans odaklƒ±"""
    try:
        output = BytesIO()
        
        # DataFrame'i kopyala ve "-" deƒüerlerini 0'a √ßevir
        df_clean = df.copy()
        
        # Depo ve tedarik√ßi bakiye kolonlarƒ±nda "-" deƒüerlerini 0'a √ßevir
        depo_cols = [col for col in df_clean.columns if any(keyword in col for keyword in 
                   ['DEVIR', 'ALI≈û', 'SATIS', 'STOK', 'Depo Bakiye', 'Tedarik√ßi Bakiye'])]
        
        for col in depo_cols:
            if col in df_clean.columns:
                # √ñnce string'e √ßevir, sonra temizlik yap
                df_clean[col] = df_clean[col].astype(str)
                df_clean[col] = df_clean[col].replace('-', '0')
                df_clean[col] = df_clean[col].replace('nan', '0')
                df_clean[col] = df_clean[col].replace('None', '0')
                
                # Sayƒ±sal deƒüerlere √ßevir
                df_clean[col] = pd.to_numeric(df_clean[col], errors='coerce').fillna(0)
        
        # Debug: Temizlenen kolonlarƒ± g√∂ster
        st.info(f"üîß Temizlenen kolonlar: {len(depo_cols)} adet")
        for col in depo_cols[:5]:  # ƒ∞lk 5 kolonu g√∂ster
            st.write(f"  - {col}")
        if len(depo_cols) > 5:
            st.write(f"  ... ve {len(depo_cols)-5} kolon daha")
        
        # Her zaman performans modu kullan - hƒ±z i√ßin
        # Excel olu≈üturma ve √∂zel format uygulama
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_clean.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Excel workbook ve worksheet'i al
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # D√ºzenlenmi≈ü √úr√ºn Kodu kolonuna √∂zel format uygula
            for col_num, col_name in enumerate(df_clean.columns, 1):
                if col_name == 'D√ºzenlenmi≈ü √úr√ºn Kodu':
                    # Bu kolon i√ßin √∂zel format: metin formatƒ±
                    for row_num in range(2, len(df_clean) + 2):  # Excel'de satƒ±r 1 ba≈ülƒ±k
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.number_format = '@'  # Metin formatƒ±
                    break
            
            # Toplam Depo Bakiye kolonuna form√ºl ekle
            toplam_depo_col = None
            depo_bakiye_cols = []
            
            # Depo bakiye kolonlarƒ±nƒ± bul
            for col_num, col_name in enumerate(df_clean.columns, 1):
                if 'Depo Bakiye' in col_name and col_name != 'Toplam Depo Bakiye':
                    depo_bakiye_cols.append(col_name)
                elif col_name == 'Toplam Depo Bakiye':
                    toplam_depo_col = col_num
            
            # Form√ºl ekle
            if toplam_depo_col and depo_bakiye_cols:
                for row_num in range(2, len(df_clean) + 2):  # Excel'de satƒ±r 1 ba≈ülƒ±k
                    cell = worksheet.cell(row=row_num, column=toplam_depo_col)
                    
                    # Form√ºl olu≈ütur: =SUM(Maslak Depo Bakiye:Bolu Depo Bakiye:ƒ∞mes Depo Bakiye:Ankara Depo Bakiye:ƒ∞kitelli Depo Bakiye)
                    formula_parts = []
                    for depo_col in depo_bakiye_cols:
                        # Kolon harfini bul
                        for col_idx, col_name in enumerate(df_clean.columns, 1):
                            if col_name == depo_col:
                                col_letter = chr(64 + col_idx)  # A=65, B=66, etc.
                                if col_idx > 26:
                                    col_letter = chr(64 + (col_idx // 26)) + chr(64 + (col_idx % 26))
                                formula_parts.append(f"{col_letter}{row_num}")
                                break
                    
                    if formula_parts:
                        formula = f"=SUM({','.join(formula_parts)})"
                        cell.value = formula
        
        output.seek(0)
        return output.getvalue()
    
    except Exception as e:
        # Hata durumunda da Excel olu≈ütur
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            
            # Excel workbook ve worksheet'i al
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # D√ºzenlenmi≈ü √úr√ºn Kodu kolonuna √∂zel format uygula
            for col_num, col_name in enumerate(df.columns, 1):
                if col_name == 'D√ºzenlenmi≈ü √úr√ºn Kodu':
                    # Bu kolon i√ßin √∂zel format: metin formatƒ±
                    for row_num in range(2, len(df) + 2):  # Excel'de satƒ±r 1 ba≈ülƒ±k
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.number_format = '@'  # Metin formatƒ±
                    break
            
            # Toplam Depo Bakiye kolonuna form√ºl ekle (hata durumunda)
            toplam_depo_col = None
            depo_bakiye_cols = []
            
            # Depo bakiye kolonlarƒ±nƒ± bul
            for col_num, col_name in enumerate(df.columns, 1):
                if 'Depo Bakiye' in col_name and col_name != 'Toplam Depo Bakiye':
                    depo_bakiye_cols.append(col_name)
                elif col_name == 'Toplam Depo Bakiye':
                    toplam_depo_col = col_num
            
            # Form√ºl ekle
            if toplam_depo_col and depo_bakiye_cols:
                for row_num in range(2, len(df) + 2):  # Excel'de satƒ±r 1 ba≈ülƒ±k
                    cell = worksheet.cell(row=row_num, column=toplam_depo_col)
                    
                    # Form√ºl olu≈ütur
                    formula_parts = []
                    for depo_col in depo_bakiye_cols:
                        # Kolon harfini bul
                        for col_idx, col_name in enumerate(df.columns, 1):
                            if col_name == depo_col:
                                col_letter = chr(64 + col_idx)  # A=65, B=66, etc.
                                if col_idx > 26:
                                    col_letter = chr(64 + (col_idx // 26)) + chr(64 + (col_idx % 26))
                                formula_parts.append(f"{col_letter}{row_num}")
                                break
                    
                    if formula_parts:
                        formula = f"=SUM({','.join(formula_parts)})"
                        cell.value = formula
        
        output.seek(0)
        return output.getvalue()

# Ana uygulama
def main():
    # Hata yakalama ve yeniden ba≈ülatma kontrol√º
    if 'kerim_restarted' not in st.session_state:
        st.session_state.kerim_restarted = False
    
    # Eƒüer sayfa yeniden ba≈ülatƒ±ldƒ±ysa
    if st.session_state.kerim_restarted:
        st.success("‚úÖ Sayfa ba≈üarƒ±yla yeniden ba≈ülatƒ±ldƒ±!")
        st.session_state.kerim_restarted = False
    
    # Dosya y√ºkleme alanƒ±
    with st.expander("üì§ ANA EXCEL DOSYASINI Y√úKLEYƒ∞N", expanded=True):
        uploaded_file = st.file_uploader(
            "Excel dosyasƒ±nƒ± se√ßin (XLSX/XLS)",
            type=['xlsx', 'xls'],
            key="main_file"
        )
    
    if uploaded_file:
        try:
            # Hƒ±zlƒ± i≈ülem akƒ±≈üƒ±
            with st.spinner("‚ö° Dosya i≈üleniyor..."):
                # 1. Hƒ±zlƒ± okuma
                df = load_data_ultra_fast(uploaded_file)
                st.success(f"‚úÖ Y√ºklendi: {len(df):,} satƒ±r | {len(df.columns)} s√ºtun")
                
                # 2. Hƒ±zlƒ± d√∂n√º≈ü√ºm
                transformed_df = transform_data_ultra_fast(df)
                st.session_state.processed_data = transformed_df
                
                # 3. Hƒ±zlƒ± Excel olu≈üturma
                if transformed_df is not None and len(transformed_df) > 0:
                    try:
                        excel_data = format_excel_ultra_fast(transformed_df)
                        st.download_button(
                            label=f"üì• D√∂n√º≈üt√ºr√ºlm√º≈ü Veriyi ƒ∞ndir ({len(transformed_df):,} satƒ±r)",
                            data=excel_data,
                            file_name=f"donusturulmus_veri_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                    except Exception as e:
                        st.error(f"Excel olu≈üturma hatasƒ±: {str(e)}")
                else:
                    st.warning("D√∂n√º≈üt√ºr√ºlecek veri bulunamadƒ±.")
        
        except Exception as e:
            st.error(f"‚ùå Hata: {str(e)}")
            st.error("üí° √á√∂z√ºm: Cache temizleyin veya sayfayƒ± yenileyin.")
            
            # Cache temizleme ve yeniden ba≈ülatma butonlarƒ±
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("üßπ Cache Temizle", type="secondary"):
                    if clear_all_caches():
                        st.success("‚úÖ Cache temizlendi!")
                        st.rerun()
                    else:
                        st.error("‚ùå Cache temizleme ba≈üarƒ±sƒ±z!")
            
            with col2:
                if st.button("üîÑ Sayfayƒ± Yeniden Ba≈ülat", type="secondary"):
                    st.session_state.kerim_restarted = True
                    st.rerun()
            
            st.stop()
    
    # 7 farklƒ± Excel ekleme kutusu - hƒ±zlƒ± y√ºkleme
    st.header("üìÇ Ek Excel Dosyalarƒ±nƒ± Y√ºkleme")
    st.write("A≈üaƒüƒ±daki 7 Excel dosyasƒ±nƒ± y√ºkleyin:")
    
    # 7 Excel dosyasƒ± y√ºkleme - tek s√ºtun
    excel1 = st.file_uploader("Schaeffler Luk", type=['xlsx', 'xls'], key="excel1")
    excel2 = st.file_uploader("ZF ƒ∞thal Bakiye", type=['xlsx', 'xls'], key="excel2")
    excel3 = st.file_uploader("Delphi Bakiye", type=['xlsx', 'xls'], key="excel3")
    excel4 = st.file_uploader("ZF Yerli Bakiye", type=['xlsx', 'xls'], key="excel4")
    excel5 = st.file_uploader("Valeo Bakiye", type=['xlsx', 'xls'], key="excel5")
    excel6 = st.file_uploader("Filtron Bakiye", type=['xlsx', 'xls'], key="excel6")
    excel7 = st.file_uploader("Mann Bakiye", type=['xlsx', 'xls'], key="excel7")
    
    # Y√ºkleme kontrol√º
    uploaded_files = {
        'excel1': excel1, 'excel2': excel2, 'excel3': excel3, 'excel4': excel4,
        'excel5': excel5, 'excel6': excel6, 'excel7': excel7
    }
    uploaded_count = sum(1 for file in uploaded_files.values() if file is not None)
    
    st.write(f"**Y√ºklenen dosya sayƒ±sƒ±:** {uploaded_count}/7")
    
    # G√ºncelle butonu
    if uploaded_count > 0:
        if st.button("üöÄ Ultra Hƒ±zlƒ± Marka E≈üle≈ütirme Yap", type="primary"):
            try:
                if st.session_state.processed_data is not None:
                    # Paralel marka e≈üle≈ütirme i≈ülemi
                    with st.spinner("‚ö° Marka e≈üle≈ütirme yapƒ±lƒ±yor..."):
                        final_df = match_brands_parallel(st.session_state.processed_data, uploaded_files)
                        st.success(f"‚úÖ Ultra hƒ±zlƒ± marka e≈üle≈ütirme tamamlandƒ±! {len(final_df)} satƒ±r i≈ülendi.")
                    
                    # Final Excel indirme butonu
                    if len(final_df) > 0:
                        try:
                            with st.spinner("‚ö° Final Excel olu≈üturuluyor..."):
                                final_excel_data = format_excel_ultra_fast(final_df)
                                st.download_button(
                                    label=f"üì• E≈üle≈ütirilmi≈ü Veriyi ƒ∞ndir ({len(final_df):,} satƒ±r)",
                                    data=final_excel_data,
                                    file_name=f"eslestirilmis_veri_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="primary"
                                )
                        except Exception as e:
                            st.error(f"Final Excel olu≈üturma hatasƒ±: {str(e)}")
                            st.error("üí° √á√∂z√ºm: Sayfayƒ± yenileyin ve tekrar deneyin.")
                else:
                    st.warning("√ñnce ana Excel dosyasƒ±nƒ± y√ºkleyin ve d√∂n√º≈üt√ºr√ºn.")
            except Exception as e:
                st.error(f"‚ùå Marka e≈üle≈ütirme hatasƒ±: {str(e)}")
                st.error("üí° √á√∂z√ºm: Cache temizleyin veya sayfayƒ± yenileyin.")
                
                # Cache temizleme ve yeniden ba≈ülatma butonlarƒ±
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("üßπ Cache Temizle", type="secondary"):
                        if clear_all_caches():
                            st.success("‚úÖ Cache temizlendi!")
                            st.rerun()
                        else:
                            st.error("‚ùå Cache temizleme ba≈üarƒ±sƒ±z!")
                
                with col2:
                    if st.button("üîÑ Sayfayƒ± Yeniden Ba≈ülat", type="secondary"):
                        st.session_state.kerim_restarted = True
                        st.rerun()
    else:
        st.info("L√ºtfen en az bir marka dosyasƒ± y√ºkleyin.")
    
    # Ana sayfaya d√∂n√º≈ü ve cache temizleme
    st.markdown("---")
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("üè† Ana Sayfaya D√∂n", type="secondary"):
            st.switch_page("Home")
    
    with col2:
        if st.button("üßπ Cache Temizle", type="secondary"):
            if clear_all_caches():
                st.success("‚úÖ Cache ba≈üarƒ±yla temizlendi!")
                st.session_state.app_restart_count += 1
                st.rerun()
            else:
                st.error("‚ùå Cache temizleme ba≈üarƒ±sƒ±z!")

# Sidebar
def sidebar():
    st.sidebar.header("‚ö° Maksimum Hƒ±z Modu")
    
    st.sidebar.success("""
    **Aktif Optimizasyonlar:**
    - Minimal dtype belirtme
    - Vekt√∂rel i≈ülemler
    - Bellek optimizasyonu
    - Paralel i≈üleme
    - Hƒ±zlƒ± Excel olu≈üturma
    """)
    
    # Cache durumu
    st.sidebar.markdown("---")
    st.sidebar.subheader("üßπ Cache Durumu")
    
    if st.sidebar.button("Cache Temizle", type="secondary"):
        if clear_all_caches():
            st.sidebar.success("‚úÖ Cache temizlendi!")
            st.rerun()
        else:
            st.sidebar.error("‚ùå Cache temizleme ba≈üarƒ±sƒ±z!")
    
    # Yeniden ba≈ülatma sayƒ±sƒ±
    restart_count = st.session_state.get('app_restart_count', 0)
    st.sidebar.info(f"üîÑ Yeniden ba≈ülatma sayƒ±sƒ±: {restart_count}")
    
    st.sidebar.header("üìã Kurallar")
    st.sidebar.write("- 0 deƒüerleri ‚Üí '-' olarak deƒüi≈ütirilir")
    st.sidebar.write("- Depo √∂nekleri yeni isimlere d√∂n√º≈üt√ºr√ºl√ºr")
    st.sidebar.write("- Kategori s√ºtunlarƒ± korunur")
    st.sidebar.write("- Vekt√∂rel i≈ülemler ile hƒ±zlandƒ±rma")
    
    st.sidebar.header("‚ÑπÔ∏è Bilgi")
    st.sidebar.write(f"Son G√ºncelleme: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}")

if __name__ == "__main__":
    sidebar()
    main() 